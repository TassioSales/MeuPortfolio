"""
Módulo para gerenciamento de carteiras de investimentos.
"""
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple, Any
from decimal import Decimal, ROUND_HALF_UP
import logging
from sqlalchemy import func, and_, or_
from ..models.ativo import Ativo, HistoricoPreco, Carteira, Operacao, TipoOperacao, StatusOperacao
from ..models.usuario import Usuario
from .. import db
from ..utils.logger import logger
from ..exceptions import (
    InsufficientFundsError,
    InvalidOperationError,
    AssetNotFoundError,
    DuplicateAssetError,
    ValidationError
)

class CarteiraService:
    """Classe para gerenciar operações relacionadas à carteira de investimentos."""
    
    @staticmethod
    def obter_carteira(usuario_id: int, ativo_id: Optional[int] = None) -> Dict:
        """
        Obtém os ativos da carteira de um usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int, optional): ID do ativo específico a ser buscado. Se None, retorna todos os ativos.
            
        Returns:
            Dict: Dicionário contendo os ativos da carteira e informações resumidas.
            
        Raises:
            AssetNotFoundError: Se o ativo especificado não for encontrado na carteira.
        """
        try:
            # Consulta os ativos na carteira do usuário
            query = Carteira.query.filter_by(usuario_id=usuario_id)
            
            if ativo_id is not None:
                query = query.filter_by(ativo_id=ativo_id)
            
            itens_carteira = query.all()
            
            if not itens_carteira and ativo_id is not None:
                raise AssetNotFoundError(f"Ativo com ID {ativo_id} não encontrado na carteira.")
            
            # Calcula o valor total da carteira e prepara os dados de cada ativo
            valor_total = Decimal('0')
            total_investido = Decimal('0')
            lucro_prejuizo_total = Decimal('0')
            
            ativos = []
            
            for item in itens_carteira:
                ativo = Ativo.query.get(item.ativo_id)
                if not ativo:
                    continue
                
                # Calcula o valor atual do ativo na carteira
                valor_atual = (ativo.preco_atual or Decimal('0')) * item.quantidade
                valor_investido = item.preco_medio * item.quantidade
                lucro_prejuizo = valor_atual - valor_investido
                
                # Atualiza os totais
                valor_total += valor_atual
                total_investido += valor_investido
                lucro_prejuizo_total += lucro_prejuizo
                
                # Prepara os dados do ativo
                ativo_info = {
                    'id': ativo.id,
                    'simbolo': ativo.symbol,
                    'nome': ativo.nome or ativo.symbol,
                    'quantidade': float(item.quantidade.quantize(Decimal('0.00000001'))),
                    'preco_medio': float(item.preco_medio.quantize(Decimal('0.00000001'))),
                    'preco_atual': float(ativo.preco_atual.quantize(Decimal('0.00000001'))) if ativo.preco_atual else None,
                    'valor_investido': float(valor_investido.quantize(Decimal('0.01'))),
                    'valor_atual': float(valor_atual.quantize(Decimal('0.01'))),
                    'lucro_prejuizo': float(lucro_prejuizo.quantize(Decimal('0.01'))),
                    'variacao_percentual': float((lucro_prejuizo / valor_investido * 100).quantize(Decimal('0.01'))) if valor_investido > 0 else 0,
                    'data_entrada': item.data_entrada.isoformat() if item.data_entrada else None,
                    'variacao_dia': float(ativo.variacao_percentual or 0)
                }
                
                ativos.append(ativo_info)
            
            # Calcula a variação percentual total
            variacao_percentual_total = 0
            if total_investido > 0:
                variacao_percentual_total = float((lucro_prejuizo_total / total_investido * 100).quantize(Decimal('0.01')))
            
            # Ordena os ativos por valor atual (maior primeiro)
            ativos_ordenados = sorted(ativos, key=lambda x: x['valor_atual'], reverse=True)
            
            # Calcula a participação percentual de cada ativo
            for ativo in ativos_ordenados:
                if valor_total > 0:
                    ativo['participacao_percentual'] = float((Decimal(str(ativo['valor_atual'])) / valor_total * 100).quantize(Decimal('0.01')))
                else:
                    ativo['participacao_percentual'] = 0
            
            return {
                'ativos': ativos_ordenados,
                'resumo': {
                    'valor_total': float(valor_total.quantize(Decimal('0.01'))),
                    'total_investido': float(total_investido.quantize(Decimal('0.01'))),
                    'lucro_prejuizo_total': float(lucro_prejuizo_total.quantize(Decimal('0.01'))),
                    'variacao_percentual_total': variacao_percentual_total,
                    'total_ativos': len(ativos_ordenados),
                    'data_atualizacao': datetime.utcnow().isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Erro ao obter carteira do usuário {usuario_id}: {str(e)}")
            raise
    
    @staticmethod
    def registrar_compra(
        usuario_id: int,
        ativo_id: int,
        quantidade: float,
        preco_unitario: float,
        data_operacao: Optional[datetime] = None,
        custos: float = 0,
        notas: str = None
    ) -> Dict:
        """
        Registra uma operação de compra de ativo na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int): ID do ativo sendo comprado.
            quantidade (float): Quantidade de ativos comprados.
            preco_unitario (float): Preço unitário pago por cada ativo.
            data_operacao (datetime, optional): Data da operação. Se None, usa a data/hora atual.
            custos (float, optional): Custos adicionais da operação. Defaults to 0.
            notas (str, optional): Notas adicionais sobre a operação.
            
        Returns:
            Dict: Dicionário com os detalhes da operação e o novo saldo.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado.
        """
        try:
            # Validação dos parâmetros
            if quantidade <= 0:
                raise ValidationError("A quantidade deve ser maior que zero.")
                
            if preco_unitario < 0:
                raise ValidationError("O preço unitário não pode ser negativo.")
                
            if custos < 0:
                raise ValidationError("Os custos não podem ser negativos.")
                
            # Verifica se o ativo existe
            ativo = Ativo.query.get(ativo_id)
            if not ativo:
                raise AssetNotFoundError(f"Ativo com ID {ativo_id} não encontrado.")
            
            # Converte para Decimal para evitar problemas de arredondamento
            quantidade = Decimal(str(quantidade))
            preco_unitario = Decimal(str(preco_unitario))
            custos = Decimal(str(custos))
            
            # Calcula o valor total da operação
            valor_total = (quantidade * preco_unitario) + custos
            
            # Data da operação (usa a data atual se não fornecida)
            data_operacao = data_operacao or datetime.utcnow()
            
            # Verifica se já existe uma posição deste ativo na carteira do usuário
            posicao = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            # Atualiza ou cria a posição na carteira
            if posicao:
                # Atualiza a posição existente (média ponderada de preço)
                novo_total_quantidade = posicao.quantidade + quantidade
                novo_valor_total = (posicao.quantidade * posicao.preco_medio) + (quantidade * preco_unitario)
                
                posicao.preco_medio = novo_valor_total / novo_total_quantidade if novo_total_quantidade > 0 else Decimal('0')
                posicao.quantidade = novo_total_quantidade
                posicao.data_atualizacao = datetime.utcnow()
                
                # Se for a primeira compra, define a data de entrada
                if not posicao.data_entrada:
                    posicao.data_entrada = data_operacao
            else:
                # Cria uma nova posição na carteira
                posicao = Carteira(
                    usuario_id=usuario_id,
                    ativo_id=ativo_id,
                    quantidade=quantidade,
                    preco_medio=preco_unitario,
                    data_entrada=data_operacao,
                    data_atualizacao=datetime.utcnow()
                )
                db.session.add(posicao)
            
            # Registra a operação no histórico
            operacao = Operacao(
                usuario_id=usuario_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.COMPRA,
                quantidade=quantidade,
                preco_unitario=preco_unitario,
                valor_total=valor_total,
                custos=custos,
                data_operacao=data_operacao,
                status=StatusOperacao.CONCLUIDA,
                notas=notas
            )
            db.session.add(operacao)
            
            # Atualiza o preço atual do ativo
            ativo.preco_atual = preco_unitario
            ativo.ultima_atualizacao = datetime.utcnow()
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            # Busca os dados atualizados da posição
            posicao_atualizada = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            return {
                'operacao': {
                    'id': operacao.id,
                    'tipo': 'compra',
                    'quantidade': float(quantidade),
                    'preco_unitario': float(preco_unitario),
                    'valor_total': float(valor_total),
                    'custos': float(custos),
                    'data_operacao': data_operacao.isoformat(),
                    'notas': notas
                },
                'posicao_atual': {
                    'quantidade': float(posicao_atualizada.quantidade),
                    'preco_medio': float(posicao_atualizada.preco_medio),
                    'valor_total': float(posicao_atualizada.quantidade * posicao_atualizada.preco_medio)
                } if posicao_atualizada else None
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao registrar compra: {str(e)}")
            raise
    
    @staticmethod
    def registrar_dividendo(
        usuario_id: int,
        ativo_id: int,
        valor_dividendo: float,
        data_pagamento: Optional[datetime] = None,
        data_com: Optional[datetime] = None,
        notas: str = None
    ) -> Dict:
        """
        Registra o recebimento de dividendos para um ativo na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int): ID do ativo que pagou o dividendo.
            valor_dividendo (float): Valor total do dividendo recebido.
            data_pagamento (datetime, optional): Data de pagamento do dividendo. Se None, usa a data/hora atual.
            data_com (datetime, optional): Data COM (com direito) do dividendo. Se None, não é registrada.
            notas (str, optional): Notas adicionais sobre o dividendo.
            
        Returns:
            Dict: Dicionário com os detalhes do dividendo registrado.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado na carteira.
        """
        try:
            # Validação dos parâmetros
            if valor_dividendo <= 0:
                raise ValidationError("O valor do dividendo deve ser maior que zero.")
                
            # Verifica se o ativo existe na carteira do usuário
            posicao = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            if not posicao:
                raise AssetNotFoundError(f"Ativo com ID {ativo_id} não encontrado na carteira.")
            
            # Converte para Decimal para evitar problemas de arredondamento
            valor_dividendo = Decimal(str(valor_dividendo))
            
            # Data de pagamento (usa a data atual se não fornecida)
            data_pagamento = data_pagamento or datetime.utcnow()
            
            # Registra o dividendo no histórico
            operacao = Operacao(
                usuario_id=usuario_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.DIVIDENDO,
                quantidade=Decimal('0'),  # Não altera a quantidade de ativos
                preco_unitario=Decimal('0'),
                valor_total=valor_dividendo,
                data_operacao=data_pagamento,
                data_com=data_com,
                status=StatusOperacao.CONCLUIDA,
                notas=notas
            )
            db.session.add(operacao)
            
            # Atualiza estatísticas do ativo
            ativo = Ativo.query.get(ativo_id)
            if ativo:
                ativo.ultimo_dividendo = valor_dividendo
                ativo.data_ultimo_dividendo = data_pagamento
                ativo.ultima_atualizacao = datetime.utcnow()
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            return {
                'id': operacao.id,
                'tipo': 'dividendo',
                'ativo_id': ativo_id,
                'valor': float(valor_dividendo),
                'data_pagamento': data_pagamento.isoformat(),
                'data_com': data_com.isoformat() if data_com else None,
                'notas': notas
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao registrar dividendo: {str(e)}")
            raise
    
    @staticmethod
    def registrar_agrupamento(
        usuario_id: int,
        ativo_id: int,
        proporcao: str,  # Formato: "5:1" (5 para 1)
        data_operacao: Optional[datetime] = None,
        notas: str = None
    ) -> Dict:
        """
        Registra um agrupamento (inverso do desdobramento) de ativo na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int): ID do ativo que sofreu o agrupamento.
            proporcao (str): Proporção do agrupamento no formato "X:Y" (ex: "1:5" para 1 por 5).
            data_operacao (datetime, optional): Data do agrupamento. Se None, usa a data/hora atual.
            notas (str, optional): Notas adicionais sobre o agrupamento.
            
        Returns:
            Dict: Dicionário com os detalhes do agrupamento e a nova posição.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado na carteira.
            InsufficientFundsError: Se a quantidade de ativos for insuficiente para o agrupamento.
        """
        try:
            # Validação da proporção
            try:
                antes, depois = map(int, proporcao.split(':'))
                if antes <= 0 or depois <= 0:
                    raise ValueError("Os valores da proporção devem ser positivos.")
            except (ValueError, AttributeError):
                raise ValidationError("A proporção deve estar no formato 'X:Y' (ex: '1:5').")
                
            # Verifica se o ativo existe na carteira do usuário
            posicao = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            if not posicao:
                raise AssetNotFoundError(f"Ativo com ID {ativo_id} não encontrado na carteira.")
            
            # Data da operação (usa a data atual se não fornecida)
            data_operacao = data_operacao or datetime.utcnow()
            
            # Calcula a nova quantidade após o agrupamento
            fator = Decimal(str(antes)) / Decimal(str(depois))
            
            # Verifica se a quantidade atual é suficiente para o agrupamento
            if posicao.quantidade < Decimal(str(antes)):
                raise InsufficientFundsError(
                    f"Quantidade insuficiente para o agrupamento. Mínimo necessário: {antes} ativos"
                )
            
            # Calcula a nova quantidade (arredondando para baixo)
            nova_quantidade = (posicao.quantidade // Decimal(str(antes))) * Decimal(str(depois))
            
            # Atualiza a posição na carteira
            posicao_anterior = {
                'quantidade': float(posicao.quantidade),
                'preco_medio': float(posicao.preco_medio)
            }
            
            posicao.quantidade = nova_quantidade
            posicao.preco_medio = (posicao.quantidade * posicao.preco_medio) / nova_quantidade if nova_quantidade > 0 else Decimal('0')
            posicao.data_atualizacao = datetime.utcnow()
            
            # Registra o agrupamento no histórico
            operacao = Operacao(
                usuario_id=usuario_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.AGRUPAMENTO,
                quantidade=nova_quantidade,
                preco_unitario=posicao.preco_medio,
                valor_total=Decimal('0'),  # Operação sem impacto financeiro direto
                data_operacao=data_operacao,
                status=StatusOperacao.CONCLUIDA,
                notas=f"Agrupamento {proporcao}" + (f" - {notas}" if notas else "")
            )
            db.session.add(operacao)
            
            # Atualiza o preço atual do ativo (ajustado pelo agrupamento)
            ativo = Ativo.query.get(ativo_id)
            if ativo and ativo.preco_atual:
                ativo.preco_atual = ativo.preco_atual * (Decimal(str(antes)) / Decimal(str(depois)))
                ativo.ultima_atualizacao = datetime.utcnow()
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            return {
                'operacao': {
                    'id': operacao.id,
                    'tipo': 'agrupamento',
                    'proporcao': proporcao,
                    'fator': float(fator),
                    'data_operacao': data_operacao.isoformat(),
                    'notas': notas
                },
                'posicao_anterior': posicao_anterior,
                'posicao_atual': {
                    'quantidade': float(posicao.quantidade),
                    'preco_medio': float(posicao.preco_medio),
                    'valor_total': float(posicao.quantidade * posicao.preco_medio)
                }
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao registrar agrupamento: {str(e)}")
            raise
    
    @staticmethod
    def transferir_ativo(
        usuario_origem_id: int,
        usuario_destino_id: int,
        ativo_id: int,
        quantidade: float,
        preco_unitario: Optional[float] = None,
        data_transferencia: Optional[datetime] = None,
        notas: str = None
    ) -> Dict:
        """
        Transfere um ativo de uma carteira para outra.
        
        Args:
            usuario_origem_id (int): ID do usuário de origem.
            usuario_destino_id (int): ID do usuário de destino.
            ativo_id (int): ID do ativo a ser transferido.
            quantidade (float): Quantidade a ser transferida.
            preco_unitario (float, optional): Preço unitário para a transferência. Se None, usa o preço médio.
            data_transferencia (datetime, optional): Data da transferência. Se None, usa a data/hora atual.
            notas (str, optional): Notas adicionais sobre a transferência.
            
        Returns:
            Dict: Dicionário com os detalhes da transferência e as posições atualizadas.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado na carteira de origem.
            InsufficientFundsError: Se não houver saldo suficiente para a transferência.
        """
        try:
            # Validação dos parâmetros
            if quantidade <= 0:
                raise ValidationError("A quantidade deve ser maior que zero.")
                
            if usuario_origem_id == usuario_destino_id:
                raise ValidationError("O usuário de origem e destino devem ser diferentes.")
            
            # Converte para Decimal para evitar problemas de arredondamento
            quantidade = Decimal(str(quantidade))
            
            # Data da transferência (usa a data atual se não fornecida)
            data_transferencia = data_transferencia or datetime.utcnow()
            
            # Verifica se o ativo existe na carteira de origem
            posicao_origem = Carteira.query.filter_by(
                usuario_id=usuario_origem_id,
                ativo_id=ativo_id
            ).first()
            
            if not posicao_origem or posicao_origem.quantidade < quantidade:
                raise InsufficientFundsError(
                    f"Saldo insuficiente para transferência. Disponível: {posicao_origem.quantidade if posicao_origem else 0}"
                )
            
            # Obtém o preço unitário (usa o preço médio se não fornecido)
            preco_unitario = Decimal(str(preco_unitario)) if preco_unitario is not None else posicao_origem.preco_medio
            
            # Verifica se o ativo já existe na carteira de destino
            posicao_destino = Carteira.query.filter_by(
                usuario_id=usuario_destino_id,
                ativo_id=ativo_id
            ).first()
            
            # Salva as posições anteriores
            posicao_origem_anterior = {
                'quantidade': float(posicao_origem.quantidade),
                'preco_medio': float(posicao_origem.preco_medio)
            }
            
            posicao_destino_anterior = {
                'quantidade': float(posicao_destino.quantidade) if posicao_destino else 0,
                'preco_medio': float(posicao_destino.preco_medio) if posicao_destino else 0
            }
            
            # Atualiza a posição de origem (remove a quantidade transferida)
            posicao_origem.quantidade -= quantidade
            posicao_origem.data_atualizacao = datetime.utcnow()
            
            # Se a quantidade zerar, remove a posição
            if posicao_origem.quantidade <= 0:
                db.session.delete(posicao_origem)
            
            # Atualiza ou cria a posição de destino
            if posicao_destino:
                # Calcula o novo preço médio (média ponderada)
                valor_total_origem = quantidade * preco_unitario
                valor_total_destino = posicao_destino.quantidade * posicao_destino.preco_medio
                nova_quantidade = posicao_destino.quantidade + quantidade
                
                if nova_quantidade > 0:
                    novo_preco_medio = (valor_total_origem + valor_total_destino) / nova_quantidade
                else:
                    novo_preco_medio = Decimal('0')
                
                posicao_destino.quantidade = nova_quantidade
                posicao_destino.preco_medio = novo_preco_medio
                posicao_destino.data_atualizacao = datetime.utcnow()
            else:
                # Cria uma nova posição na carteira de destino
                posicao_destino = Carteira(
                    usuario_id=usuario_destino_id,
                    ativo_id=ativo_id,
                    quantidade=quantidade,
                    preco_medio=preco_unitario,
                    data_entrada=data_transferencia,
                    data_atualizacao=datetime.utcnow()
                )
                db.session.add(posicao_destino)
            
            # Registra a operação de saída na origem
            operacao_saida = Operacao(
                usuario_id=usuario_origem_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.TRANSFERENCIA_SAIDA,
                quantidade=quantidade,
                preco_unitario=preco_unitario,
                valor_total=quantidade * preco_unitario,
                data_operacao=data_transferencia,
                status=StatusOperacao.CONCLUIDA,
                notas=f"Transferência para usuário {usuario_destino_id}" + (f" - {notas}" if notas else "")
            )
            db.session.add(operacao_saida)
            
            # Registra a operação de entrada no destino
            operacao_entrada = Operacao(
                usuario_id=usuario_destino_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.TRANSFERENCIA_ENTRADA,
                quantidade=quantidade,
                preco_unitario=preco_unitario,
                valor_total=quantidade * preco_unitario,
                data_operacao=data_transferencia,
                status=StatusOperacao.CONCLUIDA,
                notas=f"Transferência do usuário {usuario_origem_id}" + (f" - {notas}" if notas else "")
            )
            db.session.add(operacao_entrada)
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            # Busca as posições atualizadas
            posicao_origem_atualizada = Carteira.query.filter_by(
                usuario_id=usuario_origem_id,
                ativo_id=ativo_id
            ).first()
            
            posicao_destino_atualizada = Carteira.query.filter_by(
                usuario_id=usuario_destino_id,
                ativo_id=ativo_id
            ).first()
            
            return {
                'transferencia': {
                    'id_operacao_saida': operacao_saida.id,
                    'id_operacao_entrada': operacao_entrada.id,
                    'quantidade': float(quantidade),
                    'preco_unitario': float(preco_unitario),
                    'data_transferencia': data_transferencia.isoformat(),
                    'notas': notas
                },
                'origem': {
                    'anterior': posicao_origem_anterior,
                    'atual': {
                        'quantidade': float(posicao_origem_atualizada.quantidade) if posicao_origem_atualizada else 0,
                        'preco_medio': float(posicao_origem_atualizada.preco_medio) if posicao_origem_atualizada else 0,
                        'valor_total': float(posicao_origem_atualizada.quantidade * posicao_origem_atualizada.preco_medio) if posicao_origem_atualizada else 0
                    } if posicao_origem_atualizada else None
                },
                'destino': {
                    'anterior': posicao_destino_anterior,
                    'atual': {
                        'quantidade': float(posicao_destino_atualizada.quantidade) if posicao_destino_atualizada else 0,
                        'preco_medio': float(posicao_destino_atualizada.preco_medio) if posicao_destino_atualizada else 0,
                        'valor_total': float(posicao_destino_atualizada.quantidade * posicao_destino_atualizada.preco_medio) if posicao_destino_atualizada else 0
                    } if posicao_destino_atualizada else None
                }
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao transferir ativo: {str(e)}")
            raise
    
    @staticmethod
    def calcular_preco_medio_carteira(usuario_id: int) -> Dict:
        """
        Calcula o preço médio de todos os ativos na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            
        Returns:
            Dict: Dicionário com o preço médio total e detalhes por ativo.
        """
        try:
            # Obtém todas as posições do usuário
            posicoes = Carteira.query.filter_by(usuario_id=usuario_id).all()
            
            # Calcula o valor total e o preço médio ponderado
            valor_total = Decimal('0')
            quantidade_total = Decimal('0')
            
            detalhes_ativos = []
            
            for posicao in posicoes:
                valor_atual = posicao.quantidade * posicao.preco_medio
                valor_total += valor_atual
                quantidade_total += posicao.quantidade
                
                # Obtém informações do ativo
                ativo = Ativo.query.get(posicao.ativo_id)
                
                detalhes_ativos.append({
                    'ativo_id': posicao.ativo_id,
                    'ticker': ativo.ticker if ativo else 'N/A',
                    'nome': ativo.nome if ativo else 'Ativo desconhecido',
                    'quantidade': float(posicao.quantidade),
                    'preco_medio': float(posicao.preco_medio),
                    'valor_total': float(valor_atual),
                    'percentual_carteira': float((valor_atual / valor_total) * 100) if valor_total > 0 else 0
                })
            
            # Calcula o preço médio global
            preco_medio_global = valor_total / quantidade_total if quantidade_total > 0 else Decimal('0')
            
            return {
                'preco_medio_global': float(preco_medio_global),
                'valor_total_carteira': float(valor_total),
                'quantidade_total_ativos': float(quantidade_total),
                'total_ativos_diferentes': len(posicoes),
                'detalhes_ativos': detalhes_ativos,
                'data_calculo': datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Erro ao calcular preço médio da carteira: {str(e)}")
            raise
    
    @staticmethod
    def criar_alerta_preco(
        usuario_id: int,
        ativo_id: int,
        preco_alvo: float,
        tipo_alerta: str = 'ATINGIU',  # 'ATINGIU', 'ACIMA_DE', 'ABAIXO_DE'
        ativo: bool = True,
        notificar_email: bool = True,
        notificar_push: bool = False,
        notas: str = None
    ) -> Dict:
        """
        Cria um alerta de preço para um ativo na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int): ID do ativo a ser monitorado.
            preco_alvo (float): Preço alvo para o alerta.
            tipo_alerta (str): Tipo de alerta ('ATINGIU', 'ACIMA_DE', 'ABAIXO_DE').
            ativo (bool): Se o alerta está ativo. Padrão: True.
            notificar_email (bool): Se deve notificar por e-mail. Padrão: True.
            notificar_push (bool): Se deve notificar por push. Padrão: False.
            notas (str, optional): Notas adicionais sobre o alerta.
            
        Returns:
            Dict: Dicionário com os detalhes do alerta criado.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado.
        """
        try:
            # Validação dos parâmetros
            if preco_alvo <= 0:
                raise ValidationError("O preço alvo deve ser maior que zero.")
                
            if tipo_alerta not in ['ATINGIU', 'ACIMA_DE', 'ABAIXO_DE']:
                raise ValidationError("Tipo de alerta inválido. Use 'ATINGIU', 'ACIMA_DE' ou 'ABAIXO_DE'.")
            
            # Verifica se o ativo existe
            ativo_obj = Ativo.query.get(ativo_id)
            if not ativo_obj:
                raise AssetNotFoundError(f"Ativo com ID {ativo_id} não encontrado.")
            
            # Verifica se o usuário tem o ativo na carteira (opcional, dependendo da regra de negócio)
            posicao = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            if not posicao:
                logger.warning(f"Usuário {usuario_id} não possui o ativo {ativo_id} na carteira. Alerta será criado mesmo assim.")
            
            # Cria o alerta
            alerta = AlertaPreco(
                usuario_id=usuario_id,
                ativo_id=ativo_id,
                preco_alvo=Decimal(str(preco_alvo)),
                tipo_alerta=tipo_alerta,
                ativo=ativo,
                notificar_email=notificar_email,
                notificar_push=notificar_push,
                notas=notas,
                data_criacao=datetime.utcnow(),
                ultima_atualizacao=datetime.utcnow()
            )
            
            db.session.add(alerta)
            db.session.commit()
            
            return {
                'id': alerta.id,
                'usuario_id': alerta.usuario_id,
                'ativo_id': alerta.ativo_id,
                'ticker': ativo_obj.ticker,
                'nome_ativo': ativo_obj.nome,
                'preco_alvo': float(alerta.preco_alvo),
                'tipo_alerta': alerta.tipo_alerta,
                'preco_atual': float(ativo_obj.preco_atual) if ativo_obj.preco_atual else None,
                'ativo': alerta.ativo,
                'notificar_email': alerta.notificar_email,
                'notificar_push': alerta.notificar_push,
                'notas': alerta.notas,
                'data_criacao': alerta.data_criacao.isoformat(),
                'ultima_atualizacao': alerta.ultima_atualizacao.isoformat()
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao criar alerta de preço: {str(e)}")
            raise
    
    @staticmethod
    def verificar_alertas_preco(usuario_id: Optional[int] = None) -> List[Dict]:
        """
        Verifica todos os alertas de preço ativos e retorna os que foram acionados.
        
        Args:
            usuario_id (int, optional): ID do usuário para filtrar. Se None, verifica todos os usuários.
            
        Returns:
            List[Dict]: Lista de alertas que foram acionados.
        """
        try:
            # Consulta os alertas ativos
            query = AlertaPreco.query.filter_by(ativo=True)
            
            if usuario_id is not None:
                query = query.filter_by(usuario_id=usuario_id)
            
            alertas = query.all()
            alertas_acionados = []
            
            for alerta in alertas:
                ativo = Ativo.query.get(alerta.ativo_id)
                if not ativo or not ativo.preco_atual:
                    continue
                
                preco_atual = ativo.preco_atual
                preco_alvo = alerta.preco_alvo
                acionado = False
                
                # Verifica se o alerta foi acionado
                if alerta.tipo_alerta == 'ATINGIU' and preco_atual == preco_alvo:
                    acionado = True
                elif alerta.tipo_alerta == 'ACIMA_DE' and preco_atual > preco_alvo:
                    acionado = True
                elif alerta.tipo_alerta == 'ABAIXO_DE' and preco_atual < preco_alvo:
                    acionado = True
                
                if acionado:
                    # Registra o acionamento
                    alerta_acionado = AlertaAcionado(
                        alerta_id=alerta.id,
                        preco_alvo=preco_alvo,
                        preco_atingido=preco_atual,
                        data_acionamento=datetime.utcnow()
                    )
                    db.session.add(alerta_acionado)
                    
                    # Desativa o alerta se for do tipo 'ATINGIU'
                    if alerta.tipo_alerta == 'ATINGIU':
                        alerta.ativo = False
                    
                    # Prepara os dados para retorno
                    alerta_info = {
                        'id': alerta.id,
                        'usuario_id': alerta.usuario_id,
                        'ativo_id': alerta.ativo_id,
                        'ticker': ativo.ticker,
                        'nome_ativo': ativo.nome,
                        'tipo_alerta': alerta.tipo_alerta,
                        'preco_alvo': float(preco_alvo),
                        'preco_atingido': float(preco_atual),
                        'variacao_percentual': float(((preco_atual - preco_alvo) / preco_alvo * 100) if preco_alvo > 0 else 0),
                        'data_acionamento': datetime.utcnow().isoformat(),
                        'notificar_email': alerta.notificar_email,
                        'notificar_push': alerta.notificar_push,
                        'notas': alerta.notas
                    }
                    
                    alertas_acionados.append(alerta_info)
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            return alertas_acionados
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao verificar alertas de preço: {str(e)}")
            raise
    
    @staticmethod
    def enviar_notificacao(
        usuario_id: int,
        titulo: str,
        mensagem: str,
        tipo: str = 'info',
        dados_adicionais: Dict = None,
        prioridade: str = 'normal',
        canal: str = 'sistema'
    ) -> Dict:
        """
        Envia uma notificação para o usuário.
        
        Args:
            usuario_id (int): ID do usuário destinatário.
            titulo (str): Título da notificação.
            mensagem (str): Mensagem da notificação.
            tipo (str): Tipo da notificação ('info', 'alerta', 'erro', 'sucesso').
            dados_adicionais (Dict, optional): Dados adicionais para a notificação.
            prioridade (str, optional): Prioridade da notificação ('baixa', 'normal', 'alta').
            canal (str, optional): Canal de envio ('email', 'push', 'sistema').
            
        Returns:
            Dict: Dicionário com o resultado do envio da notificação.
        """
        try:
            # Validação dos parâmetros
            if not titulo or not mensagem:
                raise ValueError("Título e mensagem são obrigatórios.")
                
            if tipo not in ['info', 'alerta', 'erro', 'sucesso']:
                raise ValueError("Tipo de notificação inválido.")
                
            if prioridade not in ['baixa', 'normal', 'alta']:
                raise ValueError("Prioridade inválida.")
                
            # Cria a notificação no banco de dados
            notificacao = Notificacao(
                usuario_id=usuario_id,
                titulo=titulo,
                mensagem=mensagem,
                tipo=tipo,
                dados_adicionais=dados_adicionais if dados_adicionais else {},
                prioridade=prioridade,
                canal=canal,
                lida=False,
                data_criacao=datetime.utcnow()
            )
            
            db.session.add(notificacao)
            
            # Envia a notificação pelos canais configurados
            if canal == 'email':
                # Implementar lógica de envio de e-mail
                pass
            elif canal == 'push':
                # Implementar lógica de notificação push
                pass
            
            db.session.commit()
            
            return {
                'id': notificacao.id,
                'usuario_id': notificacao.usuario_id,
                'titulo': notificacao.titulo,
                'mensagem': notificacao.mensagem,
                'tipo': notificacao.tipo,
                'prioridade': notificacao.prioridade,
                'canal': notificacao.canal,
                'lida': notificacao.lida,
                'data_criacao': notificacao.data_criacao.isoformat(),
                'enviada_com_sucesso': True
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao enviar notificação: {str(e)}")
            raise
    
    @staticmethod
    def obter_notificacoes_nao_lidas(usuario_id: int, limite: int = 10) -> List[Dict]:
        """
        Obtém as notificações não lidas de um usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            limite (int, optional): Número máximo de notificações a retornar. Padrão: 10.
            
        Returns:
            List[Dict]: Lista de notificações não lidas.
        """
        try:
            notificacoes = Notificacao.query.filter_by(
                usuario_id=usuario_id,
                lida=False
            ).order_by(
                Notificacao.data_criacao.desc()
            ).limit(limite).all()
            
            return [{
                'id': n.id,
                'titulo': n.titulo,
                'mensagem': n.mensagem,
                'tipo': n.tipo,
                'prioridade': n.prioridade,
                'canal': n.canal,
                'data_criacao': n.data_criacao.isoformat(),
                'dados_adicionais': n.dados_adicionais or {}
            } for n in notificacoes]
            
        except Exception as e:
            logger.error(f"Erro ao obter notificações não lidas: {str(e)}")
            raise
    
    @staticmethod
    def marcar_notificacao_como_lida(notificacao_id: int, usuario_id: int) -> bool:
        """
        Marca uma notificação como lida.
        
        Args:
            notificacao_id (int): ID da notificação.
            usuario_id (int): ID do usuário (para validação de segurança).
            
        Returns:
            bool: True se a notificação foi marcada como lida, False caso contrário.
        """
        try:
            notificacao = Notificacao.query.filter_by(
                id=notificacao_id,
                usuario_id=usuario_id
            ).first()
            
            if not notificacao:
                return False
                
            notificacao.lida = True
            notificacao.data_leitura = datetime.utcnow()
            db.session.commit()
            
            return True
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao marcar notificação como lida: {str(e)}")
            return False
    
    @staticmethod
    def enviar_relatorio_desempenho_email(usuario_id: int, periodo: str = 'mensal') -> Dict:
        """
        Gera e envia um relatório de desempenho por e-mail para o usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            periodo (str, optional): Período do relatório ('diario', 'semanal', 'mensal', 'anual').
                
        Returns:
            Dict: Dicionário com o status do envio do relatório.
        """
        try:
            # Obtém o usuário
            usuario = Usuario.query.get(usuario_id)
            if not usuario:
                raise ValueError("Usuário não encontrado.")
                
            # Define o período do relatório
            data_fim = datetime.utcnow()
            
            if periodo == 'diario':
                data_inicio = data_fim - timedelta(days=1)
                titulo_periodo = "diário"
            elif periodo == 'semanal':
                data_inicio = data_fim - timedelta(weeks=1)
                titulo_periodo = "semanal"
            elif periodo == 'anual':
                data_inicio = data_fim - timedelta(days=365)
                titulo_periodo = "anual"
            else:  # mensal
                data_inicio = data_fim - timedelta(days=30)
                titulo_periodo = "mensal"
            
            # Gera o relatório de desempenho
            relatorio = CarteiraService.gerar_relatorio_desempenho(
                usuario_id=usuario_id,
                data_inicio=data_inicio,
                data_fim=data_fim
            )
            
            # Formata os dados para o e-mail
            dados_email = {
                'nome_usuario': usuario.nome or 'Cliente',
                'periodo': titulo_periodo,
                'data_inicio': data_inicio.strftime('%d/%m/%Y'),
                'data_fim': data_fim.strftime('%d/%m/%Y'),
                'retorno_percentual': f"{relatorio['retorno_percentual']:.2f}%",
                'retorno_absoluto': f"R$ {relatorio['retorno_absoluto']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'saldo_inicial': f"R$ {relatorio['saldo_inicial']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'saldo_final': f"R$ {relatorio['saldo_final']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'total_ativos': relatorio['estatisticas_gerais']['quantidade_ativos'],
                'ativos_lucro': relatorio['estatisticas_gerais']['ativos_lucro'],
                'ativos_prejuizo': relatorio['estatisticas_gerais']['ativos_prejuizo'],
                'maior_lucro': f"R$ {relatorio['estatisticas_gerais']['maior_lucro']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'maior_prejuizo': f"R$ {relatorio['estatisticas_gerais']['maior_prejuizo']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'top_ativos': sorted(
                    relatorio['detalhes_ativos'],
                    key=lambda x: x['lucro_absoluto'],
                    reverse=True
                )[:5],
                'piores_ativos': sorted(
                    relatorio['detalhes_ativos'],
                    key=lambda x: x['lucro_absoluto']
                )[:5],
                'distribuicao_setorial': relatorio['distribuicao_setorial'][:5]
            }
            
            # Envia o e-mail (implementação simulada)
            # Em produção, seria necessário integrar com um serviço de e-mail
            try:
                # Exemplo de como seria com Flask-Mail ou serviço similar
                # send_email(
                #     to=usuario.email,
                #     subject=f"Relatório {titulo_periodo.capitalize()} de Desempenho da Carteira",
                #     template='email/relatorio_desempenho.html',
                #     **dados_email
                # )
                logger.info(f"E-mail de relatório enviado para {usuario.email}")
                
                # Registra a notificação no sistema
                CarteiraService.enviar_notificacao(
                    usuario_id=usuario_id,
                    titulo=f"Relatório {titulo_periodo.capitalize()} Enviado",
                    mensagem=f"Seu relatório {titulo_periodo} de desempenho foi enviado para o e-mail cadastrado.",
                    tipo='sucesso',
                    canal='sistema'
                )
                
                return {
                    'sucesso': True,
                    'mensagem': f'Relatório {titulo_periodo} enviado com sucesso para {usuario.email}',
                    'email': usuario.email,
                    'periodo': titulo_periodo,
                    'data_envio': datetime.utcnow().isoformat()
                }
                
            except Exception as email_error:
                logger.error(f"Erro ao enviar e-mail: {str(email_error)}")
                raise Exception("Erro ao enviar o e-mail com o relatório.")
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório de desempenho por e-mail: {str(e)}")
            raise
    
    @staticmethod
    def obter_configuracoes_usuario(usuario_id: int) -> Dict:
        """
        Obtém as configurações do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            
        Returns:
            Dict: Dicionário com as configurações do usuário.
        """
        try:
            # Obtém as configurações do usuário
            config = ConfiguracaoUsuario.query.filter_by(usuario_id=usuario_id).first()
            
            # Se não existir, cria uma configuração padrão
            if not config:
                config = ConfiguracaoUsuario(
                    usuario_id=usuario_id,
                    moeda_padrao='BRL',
                    fuso_horario='America/Sao_Paulo',
                    notificacoes_email=True,
                    notificacoes_push=False,
                    tema='claro',
                    data_criacao=datetime.utcnow(),
                    ultima_atualizacao=datetime.utcnow()
                )
                db.session.add(config)
                db.session.commit()
            
            # Converte para dicionário
            config_dict = {
                'usuario_id': config.usuario_id,
                'moeda_padrao': config.moeda_padrao,
                'fuso_horario': config.fuso_horario,
                'notificacoes_email': config.notificacoes_email,
                'notificacoes_push': config.notificacoes_push,
                'tema': config.tema,
                'preferencias_notificacao': config.preferencias_notificacao or {},
                'data_criacao': config.data_criacao.isoformat(),
                'ultima_atualizacao': config.ultima_atualizacao.isoformat()
            }
            
            return config_dict
            
        except Exception as e:
            logger.error(f"Erro ao obter configurações do usuário: {str(e)}")
            raise
    
    @staticmethod
    def atualizar_configuracoes_usuario(usuario_id: int, **kwargs) -> Dict:
        """
        Atualiza as configurações do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            **kwargs: Configurações a serem atualizadas.
                - moeda_padrao (str, optional): Código da moeda padrão (ex: 'BRL', 'USD').
                - fuso_horario (str, optional): Fuso horário (ex: 'America/Sao_Paulo').
                - notificacoes_email (bool, optional): Se as notificações por e-mail estão ativadas.
                - notificacoes_push (bool, optional): Se as notificações push estão ativadas.
                - tema (str, optional): Tema de interface ('claro' ou 'escuro').
                - preferencias_notificacao (dict, optional): Preferências de notificação.
                
        Returns:
            Dict: Dicionário com as configurações atualizadas do usuário.
        """
        try:
            # Obtém as configurações existentes
            config = ConfiguracaoUsuario.query.filter_by(usuario_id=usuario_id).first()
            
            # Se não existir, cria uma nova configuração
            if not config:
                config = ConfiguracaoUsuario(usuario_id=usuario_id)
                db.session.add(config)
            
            # Atualiza os campos fornecidos
            campos_permitidos = [
                'moeda_padrao', 'fuso_horario', 'notificacoes_email',
                'notificacoes_push', 'tema', 'preferencias_notificacao'
            ]
            
            for campo, valor in kwargs.items():
                if campo in campos_permitidos and hasattr(config, campo):
                    setattr(config, campo, valor)
            
            config.ultima_atualizacao = datetime.utcnow()
            db.session.commit()
            
            # Retorna as configurações atualizadas
            return CarteiraService.obter_configuracoes_usuario(usuario_id)
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao atualizar configurações do usuário: {str(e)}")
            raise
    
    @staticmethod
    def exportar_dados_usuario(usuario_id: int, formato: str = 'json') -> Dict:
        """
        Exporta os dados do usuário no formato especificado.
        
        Args:
            usuario_id (int): ID do usuário.
            formato (str, optional): Formato de exportação ('json', 'csv', 'xlsx'). Padrão: 'json'.
            
        Returns:
            Dict: Dicionário com os dados exportados e metadados.
        """
        try:
            if formato not in ['json', 'csv', 'xlsx']:
                raise ValueError("Formato de exportação inválido. Use 'json', 'csv' ou 'xlsx'.")
            
            # Obtém os dados do usuário
            usuario = Usuario.query.get(usuario_id)
            if not usuario:
                raise ValueError("Usuário não encontrado.")
            
            # Obtém as carteiras do usuário
            carteiras = Carteira.query.filter_by(usuario_id=usuario_id).all()
            
            # Obtém o histórico de operações
            operacoes = Operacao.query.filter_by(usuario_id=usuario_id)\
                .order_by(Operacao.data_operacao.desc()).all()
            
            # Obtém os alertas
            alertas = AlertaPreco.query.filter_by(usuario_id=usuario_id).all()
            
            # Obtém as configurações
            config = ConfiguracaoUsuario.query.filter_by(usuario_id=usuario_id).first()
            
            # Prepara os dados para exportação
            dados = {
                'usuario': {
                    'id': usuario.id,
                    'nome': usuario.nome,
                    'email': usuario.email,
                    'data_cadastro': usuario.data_cadastro.isoformat() if usuario.data_cadastro else None,
                    'ultimo_acesso': usuario.ultimo_acesso.isoformat() if usuario.ultimo_acesso else None
                },
                'carteiras': [{
                    'id': c.id,
                    'nome': c.nome,
                    'descricao': c.descricao,
                    'data_criacao': c.data_criacao.isoformat(),
                    'ultima_atualizacao': c.ultima_atualizacao.isoformat() if c.ultima_atualizacao else None
                } for c in carteiras],
                'posicoes': [{
                    'id': p.id,
                    'carteira_id': p.carteira_id,
                    'ativo_id': p.ativo_id,
                    'ticker': p.ativo.ticker if p.ativo else 'N/A',
                    'quantidade': float(p.quantidade),
                    'preco_medio': float(p.preco_medio),
                    'valor_investido': float(p.quantidade * p.preco_medio),
                    'data_primeira_compra': p.data_primeira_compra.isoformat() if p.data_primeira_compra else None,
                    'data_ultima_atualizacao': p.data_ultima_atualizacao.isoformat() if p.data_ultima_atualizacao else None
                } for p in carteiras],
                'operacoes': [{
                    'id': o.id,
                    'tipo': o.tipo.value,
                    'ativo_id': o.ativo_id,
                    'ticker': o.ativo.ticker if o.ativo else 'N/A',
                    'quantidade': float(o.quantidade),
                    'preco_unitario': float(o.preco_unitario) if o.preco_unitario else None,
                    'valor_total': float(o.valor_total) if o.valor_total else None,
                    'data_operacao': o.data_operacao.isoformat(),
                    'taxas': float(o.taxas) if o.taxas else 0.0,
                    'notas': o.notas,
                    'carteira_origem_id': o.carteira_origem_id,
                    'carteira_destino_id': o.carteira_destino_id
                } for o in operacoes],
                'alertas': [{
                    'id': a.id,
                    'ativo_id': a.ativo_id,
                    'ticker': a.ativo.ticker if a.ativo else 'N/A',
                    'preco_alvo': float(a.preco_alvo),
                    'tipo_alerta': a.tipo_alerta,
                    'ativo': a.ativo,
                    'notificar_email': a.notificar_email,
                    'notificar_push': a.notificar_push,
                    'notas': a.notas,
                    'data_criacao': a.data_criacao.isoformat(),
                    'ultima_atualizacao': a.ultima_atualizacao.isoformat() if a.ultima_atualizacao else None
                } for a in alertas],
                'configuracoes': {
                    'moeda_padrao': config.moeda_padrao if config else 'BRL',
                    'fuso_horario': config.fuso_horario if config else 'America/Sao_Paulo',
                    'notificacoes_email': config.notificacoes_email if config else True,
                    'notificacoes_push': config.notificacoes_push if config else False,
                    'tema': config.tema if config else 'claro',
                    'preferencias_notificacao': config.preferencias_notificacao if config else {}
                },
                'metadados': {
                    'data_exportacao': datetime.utcnow().isoformat(),
                    'formato': formato,
                    'versao': '1.0'
                }
            }
            
            # Gera o conteúdo no formato solicitado
            if formato == 'json':
                conteudo = json.dumps(dados, ensure_ascii=False, indent=2)
                nome_arquivo = f'dados_usuario_{usuario_id}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.json'
            elif formato == 'csv':
                # Para CSV, criamos um arquivo ZIP com vários CSVs
                import io
                import csv
                import zipfile
                
                output = io.BytesIO()
                with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Dados do usuário
                    user_data = io.StringIO()
                    writer = csv.writer(user_data)
                    writer.writerow(['id', 'nome', 'email', 'data_cadastro', 'ultimo_acesso'])
                    writer.writerow([
                        dados['usuario']['id'],
                        dados['usuario']['nome'],
                        dados['usuario']['email'],
                        dados['usuario']['data_cadastro'],
                        dados['usuario']['ultimo_acesso']
                    ])
                    zipf.writestr('usuario.csv', user_data.getvalue())
                    
                    # Carteiras
                    if dados['carteiras']:
                        carteiras_data = io.StringIO()
                        writer = csv.writer(carteiras_data)
                        writer.writerow(['id', 'nome', 'descricao', 'data_criacao', 'ultima_atualizacao'])
                        for c in dados['carteiras']:
                            writer.writerow([c['id'], c['nome'], c['descricao'], c['data_criacao'], c['ultima_atualizacao']])
                        zipf.writestr('carteiras.csv', carteiras_data.getvalue())
                    
                    # Posições
                    if dados['posicoes']:
                        posicoes_data = io.StringIO()
                        writer = csv.writer(posicoes_data)
                        writer.writerow([
                            'id', 'carteira_id', 'ativo_id', 'ticker', 'quantidade', 
                            'preco_medio', 'valor_investido', 'data_primeira_compra', 'data_ultima_atualizacao'
                        ])
                        for p in dados['posicoes']:
                            writer.writerow([
                                p['id'], p['carteira_id'], p['ativo_id'], p['ticker'], p['quantidade'],
                                p['preco_medio'], p['valor_investido'], p['data_primeira_compra'], p['data_ultima_atualizacao']
                            ])
                        zipf.writestr('posicoes.csv', posicoes_data.getvalue())
                    
                    # Operações
                    if dados['operacoes']:
                        operacoes_data = io.StringIO()
                        writer = csv.writer(operacoes_data)
                        writer.writerow([
                            'id', 'tipo', 'ativo_id', 'ticker', 'quantidade', 'preco_unitario',
                            'valor_total', 'data_operacao', 'taxas', 'notas', 'carteira_origem_id', 'carteira_destino_id'
                        ])
                        for o in dados['operacoes']:
                            writer.writerow([
                                o['id'], o['tipo'], o['ativo_id'], o['ticker'], o['quantidade'],
                                o['preco_unitario'], o['valor_total'], o['data_operacao'], o['taxas'],
                                o['notas'], o['carteira_origem_id'], o['carteira_destino_id']
                            ])
                        zipf.writestr('operacoes.csv', operacoes_data.getvalue())
                    
                    # Alertas
                    if dados['alertas']:
                        alertas_data = io.StringIO()
                        writer = csv.writer(alertas_data)
                        writer.writerow([
                            'id', 'ativo_id', 'ticker', 'preco_alvo', 'tipo_alerta', 'ativo',
                            'notificar_email', 'notificar_push', 'notas', 'data_criacao', 'ultima_atualizacao'
                        ])
                        for a in dados['alertas']:
                            writer.writerow([
                                a['id'], a['ativo_id'], a['ticker'], a['preco_alvo'], a['tipo_alerta'],
                                a['ativo'], a['notificar_email'], a['notificar_push'], a['notas'],
                                a['data_criacao'], a['ultima_atualizacao']
                            ])
                        zipf.writestr('alertas.csv', alertas_data.getvalue())
                    
                    # Configurações
                    config_data = io.StringIO()
                    writer = csv.writer(config_data)
                    writer.writerow(['chave', 'valor'])
                    for key, value in dados['configuracoes'].items():
                        if isinstance(value, dict):
                            writer.writerow([key, json.dumps(value, ensure_ascii=False)])
                        else:
                            writer.writerow([key, value])
                    zipf.writestr('configuracoes.csv', config_data.getvalue())
                
                conteudo = output.getvalue()
                nome_arquivo = f'dados_usuario_{usuario_id}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.zip'
                
            elif formato == 'xlsx':
                # Para Excel, usamos openpyxl
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font, Alignment
                
                wb = Workbook()
                
                # Dados do usuário
                ws = wb.active
                ws.title = "Usuário"
                ws.append(['ID', 'Nome', 'E-mail', 'Data de Cadastro', 'Último Acesso'])
                ws.append([
                    dados['usuario']['id'],
                    dados['usuario']['nome'],
                    dados['usuario']['email'],
                    dados['usuario']['data_cadastro'],
                    dados['usuario']['ultimo_acesso']
                ])
                
                # Formatação do cabeçalho
                for col in range(1, 6):
                    ws.column_dimensions[get_column_letter(col)].width = 20
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Carteiras
                if dados['carteiras']:
                    ws = wb.create_sheet("Carteiras")
                    ws.append(['ID', 'Nome', 'Descrição', 'Data de Criação', 'Última Atualização'])
                    for c in dados['carteiras']:
                        ws.append([c['id'], c['nome'], c['descricao'], c['data_criacao'], c['ultima_atualizacao']])
                    
                    # Formatação
                    for col in range(1, 6):
                        ws.column_dimensions[get_column_letter(col)].width = 25
                        cell = ws.cell(row=1, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                
                # Posições
                if dados['posicoes']:
                    ws = wb.create_sheet("Posições")
                    headers = [
                        'ID', 'Carteira ID', 'Ativo ID', 'Ticker', 'Quantidade',
                        'Preço Médio', 'Valor Investido', 'Primeira Compra', 'Última Atualização'
                    ]
                    ws.append(headers)
                    for p in dados['posicoes']:
                        ws.append([
                            p['id'], p['carteira_id'], p['ativo_id'], p['ticker'], p['quantidade'],
                            p['preco_medio'], p['valor_investido'], p['data_primeira_compra'], p['data_ultima_atualizacao']
                        ])
                    
                    # Formatação
                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col)].width = 20
                        cell = ws.cell(row=1, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                
                # Operações
                if dados['operacoes']:
                    ws = wb.create_sheet("Operações")
                    headers = [
                        'ID', 'Tipo', 'Ativo ID', 'Ticker', 'Quantidade', 'Preço Unitário',
                        'Valor Total', 'Data da Operação', 'Taxas', 'Notas',
                        'Carteira Origem ID', 'Carteira Destino ID'
                    ]
                    ws.append(headers)
                    for o in dados['operacoes']:
                        ws.append([
                            o['id'], o['tipo'], o['ativo_id'], o['ticker'], o['quantidade'],
                            o['preco_unitario'], o['valor_total'], o['data_operacao'], o['taxas'],
                            o['notas'], o['carteira_origem_id'], o['carteira_destino_id']
                        ])
                    
                    # Formatação
                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col)].width = 20
                        cell = ws.cell(row=1, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                
                # Alertas
                if dados['alertas']:
                    ws = wb.create_sheet("Alertas")
                    headers = [
                        'ID', 'Ativo ID', 'Ticker', 'Preço Alvo', 'Tipo de Alerta', 'Ativo',
                        'Notificar E-mail', 'Notificar Push', 'Notas', 'Data de Criação', 'Última Atualização'
                    ]
                    ws.append(headers)
                    for a in dados['alertas']:
                        ws.append([
                            a['id'], a['ativo_id'], a['ticker'], a['preco_alvo'], a['tipo_alerta'],
                            a['ativo'], a['notificar_email'], a['notificar_push'], a['notas'],
                            a['data_criacao'], a['ultima_atualizacao']
                        ])
                    
                    # Formatação
                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col)].width = 20
                        cell = ws.cell(row=1, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                
                # Configurações
                ws = wb.create_sheet("Configurações")
                ws.append(['Chave', 'Valor'])
                for key, value in dados['configuracoes'].items():
                    if isinstance(value, dict):
                        ws.append([key, json.dumps(value, ensure_ascii=False)])
                    else:
                        ws.append([key, str(value)])
                
                # Formatação
                for col in range(1, 3):
                    ws.column_dimensions[get_column_letter(col)].width = 30
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Salva o arquivo em memória
                output = io.BytesIO()
                wb.save(output)
                conteudo = output.getvalue()
                nome_arquivo = f'dados_usuario_{usuario_id}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            # Retorna os dados para download
            return {
                'nome_arquivo': nome_arquivo,
                'conteudo': base64.b64encode(conteudo).decode('utf-8') if not isinstance(conteudo, str) else base64.b64encode(conteudo.encode('utf-8')).decode('utf-8'),
                'tipo_conteudo': 'application/octet-stream',
                'tamanho': len(conteudo) if not isinstance(conteudo, str) else len(conteudo.encode('utf-8')),
                'formato': formato,
                'data_exportacao': datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Erro ao exportar dados do usuário: {str(e)}")
            raise
    
    @staticmethod
    def importar_dados_usuario(usuario_id: int, arquivo: bytes, formato: str) -> Dict:
        """
        Importa dados para o sistema a partir de um arquivo.
        
        Args:
            usuario_id (int): ID do usuário.
            arquivo (bytes): Conteúdo binário do arquivo a ser importado.
            formato (str): Formato do arquivo ('json', 'csv', 'xlsx').
            
        Returns:
            Dict: Dicionário com o resultado da importação.
        """
        try:
            if formato not in ['json', 'csv', 'xlsx']:
                raise ValueError("Formato de importação inválido. Use 'json', 'csv' ou 'xlsx'.")
            
            resultados = {
                'total_importados': 0,
                'erros': [],
                'detalhes': {
                    'carteiras': 0,
                    'operacoes': 0,
                    'alertas': 0
                }
            }
            
            dados = None
            
            if formato == 'json':
                try:
                    dados = json.loads(arquivo.decode('utf-8'))
                except json.JSONDecodeError as e:
                    raise ValueError(f"Erro ao decodificar o arquivo JSON: {str(e)}")
                
                # Processa os dados do JSON
                if 'carteiras' in dados and isinstance(dados['carteiras'], list):
                    for carteira_data in dados['carteiras']:
                        try:
                            # Cria ou atualiza a carteira
                            carteira = Carteira.query.filter_by(
                                usuario_id=usuario_id, 
                                nome=carteira_data.get('nome')
                            ).first()
                            
                            if not carteira:
                                carteira = Carteira(
                                    usuario_id=usuario_id,
                                    nome=carteira_data.get('nome'),
                                    descricao=carteira_data.get('descricao'),
                                    data_criacao=datetime.fromisoformat(carteira_data.get('data_criacao')) if carteira_data.get('data_criacao') else datetime.utcnow(),
                                    ultima_atualizacao=datetime.utcnow()
                                )
                                db.session.add(carteira)
                                db.session.flush()
                                resultados['detalhes']['carteiras'] += 1
                            
                            # Processa as operações da carteira, se houver
                            if 'operacoes' in carteira_data and isinstance(carteira_data['operacoes'], list):
                                for operacao_data in carteira_data['operacoes']:
                                    try:
                                        # Verifica se o ativo existe
                                        ativo = Ativo.query.filter_by(ticker=operacao_data.get('ticker')).first()
                                        if not ativo:
                                            ativo = Ativo(
                                                ticker=operacao_data.get('ticker'),
                                                nome=operacao_data.get('nome_ativo', operacao_data.get('ticker')),
                                                tipo=operacao_data.get('tipo_ativo', 'AÇÃO'),
                                                setor=operacao_data.get('setor', 'OUTROS'),
                                                preco_atual=operacao_data.get('preco_atual', 0),
                                                data_atualizacao=datetime.utcnow()
                                            )
                                            db.session.add(ativo)
                                            db.session.flush()
                                        
                                        # Cria a operação
                                        operacao = Operacao(
                                            usuario_id=usuario_id,
                                            carteira_id=carteira.id,
                                            ativo_id=ativo.id,
                                            tipo=TipoOperacao[operacao_data.get('tipo')],
                                            quantidade=Decimal(str(operacao_data.get('quantidade', 0))),
                                            preco_unitario=Decimal(str(operacao_data.get('preco_unitario', 0))),
                                            valor_total=Decimal(str(operacao_data.get('valor_total', 0))),
                                            data_operacao=datetime.fromisoformat(operacao_data.get('data_operacao')) if operacao_data.get('data_operacao') else datetime.utcnow(),
                                            taxas=Decimal(str(operacao_data.get('taxas', 0))),
                                            notas=operacao_data.get('notas'),
                                            data_criacao=datetime.utcnow(),
                                            ultima_atualizacao=datetime.utcnow()
                                        )
                                        db.session.add(operacao)
                                        resultados['detalhes']['operacoes'] += 1
                                        
                                    except Exception as e:
                                        resultados['erros'].append(f"Erro ao importar operação {operacao_data.get('id')}: {str(e)}")
                                        continue
                            
                            # Processa os alertas, se houver
                            if 'alertas' in carteira_data and isinstance(carteira_data['alertas'], list):
                                for alerta_data in carteira_data['alertas']:
                                    try:
                                        # Verifica se o ativo existe
                                        ativo = Ativo.query.filter_by(ticker=alerta_data.get('ticker')).first()
                                        if not ativo:
                                            ativo = Ativo(
                                                ticker=alerta_data.get('ticker'),
                                                nome=alerta_data.get('nome_ativo', alerta_data.get('ticker')),
                                                tipo=alerta_data.get('tipo_ativo', 'AÇÃO'),
                                                setor=alerta_data.get('setor', 'OUTROS'),
                                                preco_atual=alerta_data.get('preco_atual', 0),
                                                data_atualizacao=datetime.utcnow()
                                            )
                                            db.session.add(ativo)
                                            db.session.flush()
                                        
                                        # Cria o alerta
                                        alerta = AlertaPreco(
                                            usuario_id=usuario_id,
                                            ativo_id=ativo.id,
                                            preco_alvo=Decimal(str(alerta_data.get('preco_alvo', 0))),
                                            tipo_alerta=alerta_data.get('tipo_alerta', 'PREÇO_ATINGIDO'),
                                            ativo=alerta_data.get('ativo', True),
                                            notificar_email=alerta_data.get('notificar_email', False),
                                            notificar_push=alerta_data.get('notificar_push', False),
                                            notas=alerta_data.get('notas'),
                                            data_criacao=datetime.fromisoformat(alerta_data.get('data_criacao')) if alerta_data.get('data_criacao') else datetime.utcnow(),
                                            ultima_atualizacao=datetime.utcnow()
                                        )
                                        db.session.add(alerta)
                                        resultados['detalhes']['alertas'] += 1
                                        
                                    except Exception as e:
                                        resultados['erros'].append(f"Erro ao importar alerta {alerta_data.get('id')}: {str(e)}")
                                        continue
                            
                        except Exception as e:
                            resultados['erros'].append(f"Erro ao importar carteira {carteira_data.get('nome')}: {str(e)}")
                            continue
                
                db.session.commit()
                resultados['total_importados'] = sum(resultados['detalhes'].values())
                
            elif formato in ['csv', 'xlsx']:
                # Implementar lógica para CSV e XLSX
                # Esta é uma implementação simplificada
                # Em um cenário real, você precisaria analisar o formato específico do arquivo
                # e mapear as colunas corretamente
                
                # Exemplo de como poderia ser feito para CSV:
                if formato == 'csv':
                    try:
                        import io
                        import csv
                        
                        # Lê o arquivo CSV
                        csv_file = io.StringIO(arquivo.decode('utf-8'))
                        csv_reader = csv.DictReader(csv_file)
                        
                        for row in csv_reader:
                            # Aqui você implementaria a lógica para processar cada linha do CSV
                            # e criar/atualizar os registros correspondentes no banco de dados
                            # Este é apenas um exemplo simplificado
                            pass
                        
                    except Exception as e:
                        raise ValueError(f"Erro ao processar arquivo CSV: {str(e)}")
                
                # Para XLSX, você usaria uma biblioteca como openpyxl ou pandas
                # para ler o arquivo e processar as planilhas
                
                resultados['erros'].append(f"Importação de arquivos {formato.upper()} ainda não implementada.")
            
            return resultados
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao importar dados do usuário: {str(e)}")
            raise
    
    @staticmethod
    def gerar_relatorio_avancado(usuario_id: int, parametros: Dict) -> Dict:
        """
        Gera um relatório avançado com base nos parâmetros fornecidos.
        
        Args:
            usuario_id (int): ID do usuário.
            parametros (Dict): Parâmetros para geração do relatório.
                - data_inicio (str, opcional): Data de início no formato 'YYYY-MM-DD'.
                - data_fim (str, opcional): Data de fim no formato 'YYYY-MM-DD'.
                - carteira_ids (List[int], opcional): IDs das carteiras a serem incluídas.
                - ativo_ids (List[int], opcional): IDs dos ativos a serem incluídos.
                - tipo_operacao (str, opcional): Tipo de operação ('COMPRA', 'VENDA', 'DIVIDENDO', etc.).
                - agrupar_por (str, opcional): Critério de agrupamento ('dia', 'semana', 'mes', 'ano', 'ativo', 'carteira').
                
        Returns:
            Dict: Dicionário com os dados do relatório.
        """
        try:
            # Valida os parâmetros
            data_inicio = datetime.strptime(parametros.get('data_inicio', '2000-01-01'), '%Y-%m-%d')
            data_fim = datetime.strptime(parametros.get('data_fim', datetime.utcnow().strftime('%Y-%m-%d')), '%Y-%m-%d')
            
            if data_inicio > data_fim:
                raise ValueError("A data de início não pode ser maior que a data de fim.")
            
            # Constrói a consulta base
            query = db.session.query(
                Operacao,
                Ativo.ticker,
                Carteira.nome.label('carteira_nome')
            ).join(
                Ativo, Operacao.ativo_id == Ativo.id
            ).join(
                Carteira, Operacao.carteira_id == Carteira.id
            ).filter(
                Operacao.usuario_id == usuario_id,
                Operacao.data_operacao.between(data_inicio, data_fim)
            )
            
            # Filtra por carteiras, se especificado
            if 'carteira_ids' in parametros and parametros['carteira_ids']:
                query = query.filter(Operacao.carteira_id.in_(parametros['carteira_ids']))
            
            # Filtra por ativos, se especificado
            if 'ativo_ids' in parametros and parametros['ativo_ids']:
                query = query.filter(Operacao.ativo_id.in_(parametros['ativo_ids']))
            
            # Filtra por tipo de operação, se especificado
            if 'tipo_operacao' in parametros and parametros['tipo_operacao']:
                query = query.filter(Operacao.tipo == TipoOperacao[parametros['tipo_operacao']])
            
            # Executa a consulta
            resultados = query.order_by(Operacao.data_operacao).all()
            
            # Processa os resultados com base no critério de agrupamento
            agrupar_por = parametros.get('agrupar_por', 'mes')
            dados_agrupados = {}
            
            for operacao, ticker, carteira_nome in resultados:
                chave = None
                
                if agrupar_por == 'dia':
                    chave = operacao.data_operacao.strftime('%Y-%m-%d')
                elif agrupar_por == 'semana':
                    # Formato: AAAA-SS (ano-número da semana)
                    chave = f"{operacao.data_operacao.strftime('%Y')}-{operacao.data_operacao.strftime('%U')}"
                elif agrupar_por == 'mes':
                    chave = operacao.data_operacao.strftime('%Y-%m')
                elif agrupar_por == 'ano':
                    chave = operacao.data_operacao.strftime('%Y')
                elif agrupar_por == 'ativo':
                    chave = ticker
                elif agrupar_por == 'carteira':
                    chave = carteira_nome
                else:
                    chave = 'todos'
                
                if chave not in dados_agrupados:
                    dados_agrupados[chave] = {
                        'quantidade_operacoes': 0,
                        'valor_total': Decimal('0'),
                        'quantidade_ativos': 0,
                        'tipos_operacao': set(),
                        'detalhes': []
                    }
                
                # Atualiza as estatísticas
                dados_agrupados[chave]['quantidade_operacoes'] += 1
                dados_agrupados[chave]['valor_total'] += operacao.valor_total or Decimal('0')
                dados_agrupados[chave]['tipos_operacao'].add(operacao.tipo.value)
                
                # Adiciona os detalhes da operação
                dados_agrupados[chave]['detalhes'].append({
                    'id': operacao.id,
                    'data_operacao': operacao.data_operacao.isoformat(),
                    'tipo': operacao.tipo.value,
                    'ticker': ticker,
                    'quantidade': float(operacao.quantidade) if operacao.quantidade else None,
                    'preco_unitario': float(operacao.preco_unitario) if operacao.preco_unitario else None,
                    'valor_total': float(operacao.valor_total) if operacao.valor_total else None,
                    'carteira': carteira_nome,
                    'notas': operacao.notas
                })
            
            # Calcula a quantidade de ativos únicos por grupo
            for chave in dados_agrupados:
                dados_agrupados[chave]['quantidade_ativos'] = len({
                    d['ticker'] for d in dados_agrupados[chave]['detalhes']
                })
                
                # Converte o conjunto de tipos para lista para serialização
                dados_agrupados[chave]['tipos_operacao'] = list(dados_agrupados[chave]['tipos_operacao'])
                
                # Converte Decimal para float para serialização
                dados_agrupados[chave]['valor_total'] = float(dados_agrupados[chave]['valor_total'])
            
            # Ordena os resultados
            if agrupar_por in ['dia', 'semana', 'mes', 'ano']:
                # Ordena por data
                dados_ordenados = {
                    k: dados_agrupados[k] 
                    for k in sorted(dados_agrupados.keys())
                }
            else:
                # Ordena por valor total (decrescente)
                dados_ordenados = {
                    k: v for k, v in sorted(
                        dados_agrupados.items(),
                        key=lambda x: x[1]['valor_total'],
                        reverse=True
                    )
                }
            
            # Prepara os dados para gráficos
            dados_grafico = {
                'labels': list(dados_ordenados.keys()),
                'datasets': [{
                    'label': 'Valor Total',
                    'data': [dados_ordenados[k]['valor_total'] for k in dados_ordenados],
                    'backgroundColor': 'rgba(54, 162, 235, 0.5)',
                    'borderColor': 'rgba(54, 162, 235, 1)',
                    'borderWidth': 1
                }]
            }
            
            return {
                'parametros': parametros,
                'total_grupos': len(dados_ordenados),
                'total_operacoes': sum(d['quantidade_operacoes'] for d in dados_ordenados.values()),
                'valor_total': sum(d['valor_total'] for d in dados_ordenados.values()),
                'dados_agrupados': dados_ordenados,
                'dados_grafico': dados_grafico,
                'metadados': {
                    'data_geracao': datetime.utcnow().isoformat(),
                    'usuario_id': usuario_id
                }
            }
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório avançado: {str(e)}")
            raise
    
    @staticmethod
    def agendar_relatorio_periodico(
        usuario_id: int,
        frequencia: str,
        hora: str = '09:00',
        dias_semana: List[int] = None,
        dia_mes: int = None,
        ativo: bool = True
    ) -> Dict:
        """
        Agenda um relatório periódico para ser enviado ao usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            frequencia (str): Frequência do relatório ('diario', 'semanal', 'mensal').
            hora (str, optional): Hária do envio no formato 'HH:MM'. Padrão: '09:00'.
            dias_semana (List[int], optional): Dias da semana (0-6, onde 0 é segunda-feira).
                Obrigatório se frequencia for 'semanal'.
            dia_mes (int, optional): Dia do mês (1-31). Obrigatório se frequencia for 'mensal'.
            ativo (bool, optional): Se o agendamento está ativo. Padrão: True.
            
        Returns:
            Dict: Dicionário com os detalhes do agendamento criado/atualizado.
            
        Raises:
            ValueError: Se os parâmetros forem inválidos.
            Exception: Se ocorrer um erro ao agendar o relatório.
        """
        try:
            # Valida os parâmetros
            if frequencia not in ['diario', 'semanal', 'mensal']:
                raise ValueError("Frequência inválida. Use 'diario', 'semanal' ou 'mensal'.")
                
            # Valida a hora
            try:
                hora_obj = datetime.strptime(hora, '%H:%M')
                hora_formatada = hora_obj.strftime('%H:%M')
            except ValueError:
                raise ValueError("Formato de hora inválido. Use 'HH:MM'.")
                
            # Valida os dias da semana se for semanal
            if frequencia == 'semanal':
                if not dias_semana or not all(0 <= d <= 6 for d in dias_semana):
                    raise ValueError("Para frequência semanal, informe pelo menos um dia da semana (0-6).")
            else:
                dias_semana = None
                
            # Valida o dia do mês se for mensal
            if frequencia == 'mensal':
                if not (1 <= (dia_mes or 0) <= 31):
                    raise ValueError("Para frequência mensal, informe um dia do mês entre 1 e 31.")
            else:
                dia_mes = None
            
            # Verifica se já existe um agendamento para este usuário e frequência
            agendamento = AgendamentoRelatorio.query.filter_by(
                usuario_id=usuario_id,
                frequencia=frequencia
            ).first()
            
            if agendamento:
                # Atualiza o agendamento existente
                agendamento.hora = hora_formatada
                agendamento.dias_semana = json.dumps(dias_semana) if dias_semana else None
                agendamento.dia_mes = dia_mes
                agendamento.ativo = ativo
                agendamento.ultima_execucao = None
                agendamento.proxima_execucao = calcular_proxima_execucao(
                    frequencia, 
                    hora_formatada, 
                    dias_semana, 
                    dia_mes
                )
                agendamento.ultima_atualizacao = datetime.utcnow()
            else:
                # Cria um novo agendamento
                proxima_execucao = calcular_proxima_execucao(
                    frequencia, 
                    hora_formatada, 
                    dias_semana, 
                    dia_mes
                )
                
                agendamento = AgendamentoRelatorio(
                    usuario_id=usuario_id,
                    frequencia=frequencia,
                    hora=hora_formatada,
                    dias_semana=json.dumps(dias_semana) if dias_semana else None,
                    dia_mes=dia_mes,
                    ativo=ativo,
                    proxima_execucao=proxima_execucao,
                    data_criacao=datetime.utcnow(),
                    ultima_atualizacao=datetime.utcnow()
                )
                db.session.add(agendamento)
            
            db.session.commit()
            
            # Retorna os detalhes do agendamento
            return {
                'id': agendamento.id,
                'usuario_id': agendamento.usuario_id,
                'frequencia': agendamento.frequencia,
                'hora': agendamento.hora,
                'dias_semana': json.loads(agendamento.dias_semana) if agendamento.dias_semana else None,
                'dia_mes': agendamento.dia_mes,
                'ativo': agendamento.ativo,
                'proxima_execucao': agendamento.proxima_execucao.isoformat() if agendamento.proxima_execucao else None,
                'ultima_execucao': agendamento.ultima_execucao.isoformat() if agendamento.ultima_execucao else None,
                'data_criacao': agendamento.data_criacao.isoformat(),
                'ultima_atualizacao': agendamento.ultima_atualizacao.isoformat()
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao agendar relatório periódico: {str(e)}")
            raise
    
    @staticmethod
    def obter_agendamentos_usuario(usuario_id: int, apenas_ativos: bool = True) -> List[Dict]:
        """
        Obtém os agendamentos de relatório de um usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            apenas_ativos (bool, optional): Se True, retorna apenas os agendamentos ativos. Padrão: True.
            
        Returns:
            List[Dict]: Lista de agendamentos do usuário.
        """
        try:
            query = AgendamentoRelatorio.query.filter_by(usuario_id=usuario_id)
            
            if apenas_ativos:
                query = query.filter_by(ativo=True)
                
            agendamentos = query.order_by(AgendamentoRelatorio.frequencia).all()
            
            return [{
                'id': a.id,
                'frequencia': a.frequencia,
                'hora': a.hora,
                'dias_semana': json.loads(a.dias_semana) if a.dias_semana else None,
                'dia_mes': a.dia_mes,
                'ativo': a.ativo,
                'proxima_execucao': a.proxima_execucao.isoformat() if a.proxima_execucao else None,
                'ultima_execucao': a.ultima_execucao.isoformat() if a.ultima_execucao else None,
                'data_criacao': a.data_criacao.isoformat(),
                'ultima_atualizacao': a.ultima_atualizacao.isoformat()
            } for a in agendamentos]
            
        except Exception as e:
            logger.error(f"Erro ao obter agendamentos do usuário {usuario_id}: {str(e)}")
            raise
    
    @staticmethod
    def remover_agendamento(agendamento_id: int, usuario_id: int) -> bool:
        """
        Remove um agendamento de relatório.
        
        Args:
            agendamento_id (int): ID do agendamento a ser removido.
            usuario_id (int): ID do usuário dono do agendamento.
            
        Returns:
            bool: True se o agendamento foi removido com sucesso, False caso contrário.
            
        Raises:
            ValueError: Se o agendamento não for encontrado ou não pertencer ao usuário.
        """
        try:
            agendamento = AgendamentoRelatorio.query.filter_by(
                id=agendamento_id,
                usuario_id=usuario_id
            ).first()
            
            if not agendamento:
                raise ValueError("Agendamento não encontrado ou você não tem permissão para removê-lo.")
            
            db.session.delete(agendamento)
            db.session.commit()
            
            return True
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao remover agendamento {agendamento_id}: {str(e)}")
            raise
    
    @staticmethod
    def obter_metricas_avancadas(usuario_id: int, carteira_id: int = None) -> Dict:
        """
        Obtém métricas avançadas sobre as carteiras e operações do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            carteira_id (int, optional): ID da carteira para filtrar as métricas.
                Se None, considera todas as carteiras do usuário.
                
        Returns:
            Dict: Dicionário com as métricas calculadas.
        """
        try:
            # Consulta base para operações
            query = db.session.query(
                Operacao,
                Ativo.ticker,
                Carteira.nome.label('carteira_nome')
            ).join(
                Ativo, Operacao.ativo_id == Ativo.id
            ).join(
                Carteira, Operacao.carteira_id == Carteira.id
            ).filter(
                Operacao.usuario_id == usuario_id
            )
            
            # Filtra por carteira, se especificado
            if carteira_id:
                query = query.filter(Operacao.carteira_id == carteira_id)
            
            # Executa a consulta
            operacoes = query.order_by(Operacao.data_operacao).all()
            
            if not operacoes:
                return {
                    'mensagem': 'Nenhuma operação encontrada para os critérios fornecidos.',
                    'dados': {}
                }
            
            # Inicializa as estruturas de dados para as métricas
            metricas = {
                'geral': {
                    'total_operacoes': 0,
                    'total_investido': Decimal('0'),
                    'total_retirado': Decimal('0'),
                    'lucro_prejuizo': Decimal('0'),
                    'retorno_percentual': Decimal('0'),
                    'ticket_medio_operacao': Decimal('0'),
                    'operacoes_por_mes': {},
                    'meses_ativos': set(),
                    'estatisticas_por_tipo': {},
                    'ativos_operados': set(),
                    'carteiras_utilizadas': set()
                },
                'por_ano': {},
                'por_ativo': {},
                'por_carteira': {},
                'por_tipo_operacao': {},
                'evolucao_patrimonio': []
            }
            
            # Variáveis para cálculo de patrimônio
            saldo_acumulado = Decimal('0')
            primeiro_ano = None
            ultimo_ano = None
            
            # Processa cada operação
            for operacao, ticker, carteira_nome in operacoes:
                # Atualiza métricas gerais
                metricas['geral']['total_operacoes'] += 1
                
                # Atualiza totais por tipo de operação
                if operacao.tipo in [TipoOperacao.COMPRA, TipoOperacao.TRANSFERENCIA_ENTRADA]:
                    metricas['geral']['total_investido'] += operacao.valor_total or Decimal('0')
                    saldo_acumulado += operacao.valor_total or Decimal('0')
                elif operacao.tipo in [TipoOperacao.VENDA, TipoOperacao.TRANSFERENCIA_SAIDA]:
                    metricas['geral']['total_retirado'] += operacao.valor_total or Decimal('0')
                    saldo_acumulado -= operacao.valor_total or Decimal('0')
                
                # Atualiza métricas por ano
                ano = operacao.data_operacao.year
                if ano not in metricas['por_ano']:
                    metricas['por_ano'][ano] = {
                        'total_operacoes': 0,
                        'total_investido': Decimal('0'),
                        'total_retirado': Decimal('0'),
                        'meses_ativos': set()
                    }
                
                metricas['por_ano'][ano]['total_operacoes'] += 1
                
                if operacao.tipo in [TipoOperacao.COMPRA, TipoOperacao.TRANSFERENCIA_ENTRADA]:
                    metricas['por_ano'][ano]['total_investido'] += operacao.valor_total or Decimal('0')
                elif operacao.tipo in [TipoOperacao.VENDA, TipoOperacao.TRANSFERENCIA_SAIDA]:
                    metricas['por_ano'][ano]['total_retirado'] += operacao.valor_total or Decimal('0')
                
                mes_ano = operacao.data_operacao.strftime('%Y-%m')
                metricas['por_ano'][ano]['meses_ativos'].add(mes_ano)
                metricas['geral']['meses_ativos'].add(mes_ano)
                
                # Atualiza métricas por ativo
                if ticker not in metricas['por_ativo']:
                    metricas['por_ativo'][ticker] = {
                        'total_operacoes': 0,
                        'total_comprado': Decimal('0'),
                        'total_vendido': Decimal('0'),
                        'saldo': Decimal('0'),
                        'primeira_operacao': operacao.data_operacao,
                        'ultima_operacao': operacao.data_operacao
                    }
                else:
                    if operacao.data_operacao < metricas['por_ativo'][ticker]['primeira_operacao']:
                        metricas['por_ativo'][ticker]['primeira_operacao'] = operacao.data_operacao
                    if operacao.data_operacao > metricas['por_ativo'][ticker]['ultima_operacao']:
                        metricas['por_ativo'][ticker]['ultima_operacao'] = operacao.data_operacao
                
                metricas['por_ativo'][ticker]['total_operacoes'] += 1
                
                if operacao.tipo == TipoOperacao.COMPRA:
                    metricas['por_ativo'][ticker]['total_comprado'] += operacao.valor_total or Decimal('0')
                    metricas['por_ativo'][ticker]['saldo'] += operacao.valor_total or Decimal('0')
                elif operacao.tipo == TipoOperacao.VENDA:
                    metricas['por_ativo'][ticker]['total_vendido'] += operacao.valor_total or Decimal('0')
                    metricas['por_ativo'][ticker]['saldo'] -= operacao.valor_total or Decimal('0')
                
                # Atualiza métricas por carteira
                if carteira_nome not in metricas['por_carteira']:
                    metricas['por_carteira'][carteira_nome] = {
                        'total_operacoes': 0,
                        'total_investido': Decimal('0'),
                        'total_retirado': Decimal('0'),
                        'saldo': Decimal('0')
                    }
                
                metricas['por_carteira'][carteira_nome]['total_operacoes'] += 1
                
                if operacao.tipo in [TipoOperacao.COMPRA, TipoOperacao.TRANSFERENCIA_ENTRADA]:
                    metricas['por_carteira'][carteira_nome]['total_investido'] += operacao.valor_total or Decimal('0')
                    metricas['por_carteira'][carteira_nome]['saldo'] += operacao.valor_total or Decimal('0')
                elif operacao.tipo in [TipoOperacao.VENDA, TipoOperacao.TRANSFERENCIA_SAIDA]:
                    metricas['por_carteira'][carteira_nome]['total_retirado'] += operacao.valor_total or Decimal('0')
                    metricas['por_carteira'][carteira_nome]['saldo'] -= operacao.valor_total or Decimal('0')
                
                # Atualiza métricas por tipo de operação
                tipo_op = operacao.tipo.value
                if tipo_op not in metricas['por_tipo_operacao']:
                    metricas['por_tipo_operacao'][tipo_op] = {
                        'total_operacoes': 0,
                        'valor_total': Decimal('0'),
                        'ticket_medio': Decimal('0')
                    }
                
                metricas['por_tipo_operacao'][tipo_op]['total_operacoes'] += 1
                metricas['por_tipo_operacao'][tipo_op]['valor_total'] += operacao.valor_total or Decimal('0')
                
                # Atualiza conjuntos para contagens únicas
                metricas['geral']['ativos_operados'].add(ticker)
                metricas['geral']['carteiras_utilizadas'].add(carteira_nome)
                
                # Atualiza primeiro e último ano para cálculo de evolução
                if primeiro_ano is None or operacao.data_operacao.year < primeiro_ano:
                    primeiro_ano = operacao.data_operacao.year
                if ultimo_ano is None or operacao.data_operacao.year > ultimo_ano:
                    ultimo_ano = operacao.data_operacao.year
            
            # Calcula métricas derivadas
            metricas['geral']['lucro_prejuizo'] = (
                metricas['geral']['total_retirado'] - metricas['geral']['total_investido']
            )
            
            if metricas['geral']['total_investido'] > 0:
                metricas['geral']['retorno_percentual'] = (
                    (metricas['geral']['lucro_prejuizo'] / metricas['geral']['total_investido']) * 100
                )
            
            if metricas['geral']['total_operacoes'] > 0:
                metricas['geral']['ticket_medio_operacao'] = (
                    (metricas['geral']['total_investido'] + metricas['geral']['total_retirado']) / 
                    metricas['geral']['total_operacoes']
                )
            
            # Calcula ticket médio por tipo de operação
            for tipo_op, dados in metricas['por_tipo_operacao'].items():
                if dados['total_operacoes'] > 0:
                    dados['ticket_medio'] = dados['valor_total'] / dados['total_operacoes']
            
            # Calcula a evolução do patrimônio ao longo do tempo
            if primeiro_ano is not None and ultimo_ano is not None:
                for ano in range(primeiro_ano, ultimo_ano + 1):
                    for mes in range(1, 13):
                        mes_ano = f"{ano}-{mes:02d}"
                        # Aqui você implementaria a lógica para calcular o patrimônio em cada mês
                        # Esta é uma implementação simplificada
                        pass
            
            # Converte Decimal para float para serialização
            for key in ['total_investido', 'total_retirado', 'lucro_prejuizo', 'retorno_percentual', 'ticket_medio_operacao']:
                if key in metricas['geral']:
                    metricas['geral'][key] = float(metricas['geral'][key])
            
            for ano in metricas['por_ano']:
                for key in ['total_investido', 'total_retirado']:
                    metricas['por_ano'][ano][key] = float(metricas['por_ano'][ano][key])
                metricas['por_ano'][ano]['meses_ativos'] = sorted(list(metricas['por_ano'][ano]['meses_ativos']))
            
            for ativo in metricas['por_ativo'].values():
                for key in ['total_comprado', 'total_vendido', 'saldo']:
                    ativo[key] = float(ativo[key])
                ativo['primeira_operacao'] = ativo['primeira_operacao'].isoformat()
                ativo['ultima_operacao'] = ativo['ultima_operacao'].isoformat()
            
            for carteira in metricas['por_carteira'].values():
                for key in ['total_investido', 'total_retirado', 'saldo']:
                    carteira[key] = float(carteira[key])
            
            for tipo_op in metricas['por_tipo_operacao'].values():
                for key in ['valor_total', 'ticket_medio']:
                    tipo_op[key] = float(tipo_op[key])
            
            # Adiciona contagens únicas
            metricas['geral']['total_ativos_unicos'] = len(metricas['geral']['ativos_operados'])
            metricas['geral']['total_carteiras_utilizadas'] = len(metricas['geral']['carteiras_utilizadas'])
            metricas['geral']['total_meses_ativos'] = len(metricas['geral']['meses_ativos'])
            
            # Remove conjuntos para serialização
            for key in ['ativos_operados', 'carteiras_utilizadas', 'meses_ativos']:
                if key in metricas['geral']:
                    metricas['geral'][key] = list(metricas['geral'][key])
            
            return {
                'mensagem': 'Métricas calculadas com sucesso.',
                'dados': metricas,
                'metadados': {
                    'data_geracao': datetime.utcnow().isoformat(),
                    'usuario_id': usuario_id,
                    'carteira_id': carteira_id
                }
            }
            
        except Exception as e:
            logger.error(f"Erro ao calcular métricas avançadas: {str(e)}")
            raise
    
    @staticmethod
    def calcular_proxima_execucao(
        frequencia: str,
        hora: str,
        dias_semana: List[int] = None,
        dia_mes: int = None
    ) -> datetime:
        """
        Calcula a próxima data de execução com base na frequência e configurações.
        
        Args:
            frequencia (str): Frequência do agendamento ('diario', 'semanal', 'mensal').
            hora (str): Hora do agendamento no formato 'HH:MM'.
            dias_semana (List[int], optional): Dias da semana (0-6, onde 0 é segunda-feira).
            dia_mes (int, optional): Dia do mês (1-31).
                
        Returns:
            datetime: Próxima data de execução calculada.
        """
        try:
            agora = datetime.utcnow()
            hora_parts = list(map(int, hora.split(':')))
            
            if frequencia == 'diario':
                # Próxima execução é hoje ou amanhã na hora especificada
                proxima = agora.replace(
                    hour=hora_parts[0],
                    minute=hora_parts[1],
                    second=0,
                    microsecond=0
                )
                
                if proxima <= agora:
                    proxima += timedelta(days=1)
                    
            elif frequencia == 'semanal':
                if not dias_semana:
                    raise ValueError("Dias da semana são obrigatórios para frequência semanal.")
                
                # Encontra o próximo dia da semana a partir de hoje
                dia_atual = agora.weekday()  # 0 = segunda, 6 = domingo
                dias_para_proximo = None
                
                # Ordena os dias da semana para garantir a ordem correta
                dias_semana_ordenados = sorted(dias_semana)
                
                # Encontra o próximo dia da semana
                for dia in dias_semana_ordenados:
                    if dia > dia_atual:
                        dias_para_proximo = dia - dia_atual
                        break
                
                # Se não encontrou um dia maior que hoje, pega o primeiro dia da próxima semana
                if dias_para_proximo is None:
                    dias_para_proximo = (7 - dia_atual) + dias_semana_ordenados[0]
                
                proxima = (agora + timedelta(days=dias_para_proximo)).replace(
                    hour=hora_parts[0],
                    minute=hora_parts[1],
                    second=0,
                    microsecond=0
                )
                
            elif frequencia == 'mensal':
                if not dia_mes:
                    raise ValueError("Dia do mês é obrigatório para frequência mensal.")
                
                # Encontra o próximo dia do mês
                dia_atual = agora.day
                mes_atual = agora.month
                ano_atual = agora.year
                
                if dia_mes > dia_atual:
                    # O dia ainda não passou este mês
                    proxima = datetime(
                        year=ano_atual,
                        month=mes_atual,
                        day=min(dia_mes, 28),  # Evita problemas com fevereiro
                        hour=hora_parts[0],
                        minute=hora_parts[1],
                        second=0,
                        microsecond=0
                    )
                    
                    # Ajusta para o último dia do mês se o dia for inválido
                    ultimo_dia_mes = (datetime(ano_atual, mes_atual % 12 + 1, 1) - timedelta(days=1)).day
                    if dia_mes > ultimo_dia_mes:
                        proxima = proxima.replace(day=ultimo_dia_mes)
                    
                    if proxima <= agora:
                        # Se já passou a hora de hoje, vai para o próximo mês
                        mes_atual += 1
                        if mes_atual > 12:
                            mes_atual = 1
                            ano_atual += 1
                else:
                    # O dia já passou, vai para o próximo mês
                    mes_atual += 1
                    if mes_atual > 12:
                        mes_atual = 1
                        ano_atual += 1
                
                proxima = datetime(
                    year=ano_atual,
                    month=mes_atual,
                    day=min(dia_mes, 28),  # Evita problemas com fevereiro
                    hour=hora_parts[0],
                    minute=hora_parts[1],
                    second=0,
                    microsecond=0
                )
                
                # Ajusta para o último dia do mês se o dia for inválido
                ultimo_dia_mes = (datetime(ano_atual, mes_atual % 12 + 1, 1) - timedelta(days=1)).day
                if dia_mes > ultimo_dia_mes:
                    proxima = proxima.replace(day=ultimo_dia_mes)
            else:
                raise ValueError(f"Frequência inválida: {frequencia}")
            
            return proxima
            
        except Exception as e:
            logger.error(f"Erro ao calcular próxima execução: {str(e)}")
            raise
        """
        Agenda o envio periódico de relatórios de desempenho.
        
        Args:
            usuario_id (int): ID do usuário.
            frequencia (str): Frequência do relatório ('diario', 'semanal', 'mensal').
            hora (str, optional): Horário do envio no formato 'HH:MM'. Padrão: '09:00'.
            dias_semana (List[int], optional): Dias da semana (0=segunda a 6=domingo) para frequência semanal.
            dia_mes (int, optional): Dia do mês (1-31) para frequência mensal.
            ativo (bool, optional): Se o agendamento está ativo. Padrão: True.
            
        Returns:
            Dict: Dicionário com os detalhes do agendamento.
        """
        try:
            # Validação dos parâmetros
            if frequencia not in ['diario', 'semanal', 'mensal']:
                raise ValueError("Frequência inválida. Use 'diario', 'semanal' ou 'mensal'.")
                
            # Valida o horário
            try:
                hora_obj = datetime.strptime(hora, '%H:%M')
                hora_formatada = hora_obj.strftime('%H:%M')
            except ValueError:
                raise ValueError("Formato de hora inválido. Use 'HH:MM'.")
                
            # Valida os dias da semana para frequência semanal
            if frequencia == 'semanal':
                if not dias_semana or not all(0 <= dia <= 6 for dia in dias_semana):
                    raise ValueError("Para frequência semanal, informe os dias da semana (0-6, onde 0=segunda e 6=domingo).")
                configuracao = {'dias_semana': dias_semana}
            # Valida o dia do mês para frequência mensal
            elif frequencia == 'mensal':
                if not (1 <= dia_mes <= 31):
                    raise ValueError("Para frequência mensal, informe um dia do mês entre 1 e 31.")
                configuracao = {'dia_mes': dia_mes}
            else:
                configuracao = {}
            
            # Verifica se já existe um agendamento para este usuário e frequência
            agendamento = AgendamentoRelatorio.query.filter_by(
                usuario_id=usuario_id,
                tipo='relatorio_desempenho',
                frequencia=frequencia
            ).first()
            
            if agendamento:
                # Atualiza o agendamento existente
                agendamento.hora = hora_formatada
                agendamento.configuracao = configuracao
                agendamento.ativo = ativo
                agendamento.ultima_atualizacao = datetime.utcnow()
                acao = 'atualizado'
            else:
                # Cria um novo agendamento
                agendamento = AgendamentoRelatorio(
                    usuario_id=usuario_id,
                    tipo='relatorio_desempenho',
                    frequencia=frequencia,
                    hora=hora_formatada,
                    configuracao=configuracao,
                    ativo=ativo,
                    data_criacao=datetime.utcnow(),
                    ultima_atualizacao=datetime.utcnow()
                )
                db.session.add(agendamento)
                acao = 'criado'
            
            db.session.commit()
            
            # Se for um agendamento ativo, agenda a próxima execução
            if ativo:
                # Aqui seria necessário integrar com um sistema de agendamento como Celery Beat ou APScheduler
                # Por exemplo:
                # agendar_tarefa(
                #     'enviar_relatorio_periodico',
                #     args=[usuario_id],
                #     **obter_parametros_agendamento(frequencia, hora_formatada, configuracao)
                # )
                pass
            
            return {
                'id': agendamento.id,
                'usuario_id': agendamento.usuario_id,
                'tipo': agendamento.tipo,
                'frequencia': agendamento.frequencia,
                'hora': agendamento.hora,
                'configuracao': agendamento.configuracao,
                'ativo': agendamento.ativo,
                'proxima_execucao': None,  # Seria calculado com base na frequência
                'acao': acao,
                'mensagem': f"Agendamento {acao} com sucesso."
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao agendar relatório periódico: {str(e)}")
            raise
    
    @staticmethod
    def analise_tecnica_ativos_carteira(
        usuario_id: int,
        indicadores: List[str] = None,
        timeframe: str = '1d',
        periodo: int = 30
    ) -> Dict:
        """
        Realiza análise técnica nos ativos da carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            indicadores (List[str], optional): Lista de indicadores a serem calculados.
                Padrão: ['media_movel', 'rsi', 'macd', 'bollinger']
            timeframe (str, optional): Intervalo de tempo para análise. Padrão: '1d' (diário).
            periodo (int, optional): Número de períodos para análise. Padrão: 30.
            
        Returns:
            Dict: Dicionário com os resultados da análise técnica por ativo.
        """
        try:
            # Define os indicadores padrão se não fornecidos
            if not indicadores:
                indicadores = ['media_movel', 'rsi', 'macd', 'bollinger']
            
            # Obtém os ativos da carteira do usuário
            posicoes = Carteira.query.filter_by(usuario_id=usuario_id).all()
            
            if not posicoes:
                return {'mensagem': 'Nenhum ativo encontrado na carteira.'}
            
            resultados = {}
            
            for posicao in posicoes:
                ativo = Ativo.query.get(posicao.ativo_id)
                if not ativo:
                    continue
                
                # Obtém os dados históricos do ativo
                # (implementação simplificada - pode ser integrada com yfinance ou outra fonte de dados)
                try:
                    # Exemplo de como seria com yfinance
                    # dados_historicos = yf.download(
                    #     f"{ativo.ticker}.SA",
                    #     start=datetime.now() - timedelta(days=periodo * 2),  # Pega mais dados para cálculos
                    #     end=datetime.now(),
                    #     interval=timeframe,
                    #     progress=False
                    # )
                    
                    # Simulação de dados históricos (remover em produção)
                    dados_historicos = pd.DataFrame({
                        'Close': [100 + i + random.uniform(-5, 5) for i in range(periodo * 2)],
                        'High': [100 + i + random.uniform(0, 5) for i in range(periodo * 2)],
                        'Low': [100 + i - random.uniform(0, 5) for i in range(periodo * 2)],
                        'Volume': [random.randint(1000, 10000) for _ in range(periodo * 2)]
                    }, index=pd.date_range(end=datetime.now(), periods=periodo * 2))
                    
                    analise = {}
                    
                    # Médias Móveis
                    if 'media_movel' in indicadores:
                        dados_historicos['MA20'] = dados_historicos['Close'].rolling(window=20).mean()
                        dados_historicos['MA50'] = dados_historicos['Close'].rolling(window=50).mean()
                        dados_historicos['MA200'] = dados_historicos['Close'].rolling(window=200).mean()
                        
                        ultimo = dados_historicos.iloc[-1]
                        
                        analise['media_movel'] = {
                            'MA20': float(ultimo['MA20']),
                            'MA50': float(ultimo['MA50']),
                            'MA200': float(ultuno['MA200']),
                            'tendencia': 'alta' if ultimo['MA20'] > ultimo['MA50'] > ultimo['MA200'] else 'baixa',
                            'cruzamento_ouro': ultimo['MA20'] > ultimo['MA50'] and dados_historicos.iloc[-2]['MA20'] <= dados_historicos.iloc[-2]['MA50'],
                            'cruzamento_morte': ultimo['MA20'] < ultimo['MA50'] and dados_historicos.iloc[-2]['MA20'] >= dados_historicos.iloc[-2]['MA50']
                        }
                    
                    # RSI (Índice de Força Relativa)
                    if 'rsi' in indicadores:
                        delta = dados_historicos['Close'].diff()
                        ganho = (delta.where(delta > 0, 0)).rolling(window=14).mean()
                        perda = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
                        rs = ganho / perda
                        rsi = 100 - (100 / (1 + rs))
                        
                        analise['rsi'] = {
                            'valor': float(rsi.iloc[-1]),
                            'sobrecomprado': rsi.iloc[-1] > 70,
                            'sobrevendido': rsi.iloc[-1] < 30,
                            'tendencia': 'alta' if rsi.iloc[-1] > 50 else 'baixa'
                        }
                    
                    # MACD (Média Móvel de Convergência/Divergência)
                    if 'macd' in indicadores:
                        exp1 = dados_historicos['Close'].ewm(span=12, adjust=False).mean()
                        exp2 = dados_historicos['Close'].ewm(span=26, adjust=False).mean()
                        macd = exp1 - exp2
                        sinal = macd.ewm(span=9, adjust=False).mean()
                        
                        analise['macd'] = {
                            'valor': float(macd.iloc[-1]),
                            'sinal': float(sinal.iloc[-1]),
                            'histograma': float(macd.iloc[-1] - sinal.iloc[-1]),
                            'cruzamento_cima': macd.iloc[-1] > sinal.iloc[-1] and macd.iloc[-2] <= sinal.iloc[-2],
                            'cruzamento_baixo': macd.iloc[-1] < sinal.iloc[-1] and macd.iloc[-2] >= sinal.iloc[-2]
                        }
                    
                    # Bandas de Bollinger
                    if 'bollinger' in indicadores:
                        media = dados_historicos['Close'].rolling(window=20).mean()
                        desvio_padrao = dados_historicos['Close'].rolling(window=20).std()
                        
                        bollinger_superior = media + (desvio_padrao * 2)
                        bollinger_inferior = media - (desvio_padrao * 2)
                        
                        preco_atual = dados_historicos['Close'].iloc[-1]
                        
                        analise['bollinger'] = {
                            'media': float(media.iloc[-1]),
                            'superior': float(bollinger_superior.iloc[-1]),
                            'inferior': float(bollinger_inferior.iloc[-1]),
                            'preco_atual': float(preco_atual),
                            'percentual_banda': float((preco_atual - bollinger_inferior.iloc[-1]) / (bollinger_superior.iloc[-1] - bollinger_inferior.iloc[-1]) * 100),
                            'tocou_banda_superior': preco_atual >= bollinger_superior.iloc[-1],
                            'tocou_banda_inferior': preco_atual <= bollinger_inferior.iloc[-1]
                        }
                    
                    # Volume
                    if 'volume' in indicadores:
                        media_volume = dados_historicos['Volume'].rolling(window=20).mean()
                        volume_atual = dados_historicos['Volume'].iloc[-1]
                        
                        analise['volume'] = {
                            'atual': int(volume_atual),
                            'media_20': float(media_volume.iloc[-1]),
                            'acima_media': volume_atual > media_volume.iloc[-1] * 1.5,
                            'variacao_percentual': float(((volume_atual - media_volume.iloc[-1]) / media_volume.iloc[-1] * 100) if media_volume.iloc[-1] > 0 else 0)
                        }
                    
                    # Adiciona os resultados para este ativo
                    resultados[ativo.ticker] = {
                        'nome': ativo.nome,
                        'preco_atual': float(dados_historicos['Close'].iloc[-1]),
                        'variacao_periodo': float((dados_historicos['Close'].iloc[-1] / dados_historicos['Close'].iloc[0] - 1) * 100),
                        'analise': analise,
                        'posicao_usuario': {
                            'quantidade': float(posicao.quantidade),
                            'preco_medio': float(posicao.preco_medio),
                            'valor_investido': float(posicao.quantidade * posicao.preco_medio),
                            'valor_atual': float(posicao.quantidade * dados_historicos['Close'].iloc[-1]),
                            'lucro_prejuizo': float(posicao.quantidade * (dados_historicos['Close'].iloc[-1] - posicao.preco_medio)),
                            'lucro_percentual': float(((dados_historicos['Close'].iloc[-1] / posicao.preco_medio) - 1) * 100) if posicao.preco_medio > 0 else 0
                        }
                    }
                    
                except Exception as e:
                    logger.error(f"Erro ao analisar ativo {ativo.ticker}: {str(e)}")
                    continue
            
            return {
                'data_analise': datetime.utcnow().isoformat(),
                'timeframe': timeframe,
                'periodo': periodo,
                'indicadores_utilizados': indicadores,
                'resultados': resultados
            }
            
        except Exception as e:
            logger.error(f"Erro na análise técnica da carteira: {str(e)}")
            raise
    
    @staticmethod
    def gerar_relatorio_desempenho(
        usuario_id: int,
        data_inicio: Optional[datetime] = None,
        data_fim: Optional[datetime] = None
    ) -> Dict:
        """
        Gera um relatório de desempenho da carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            data_inicio (datetime, optional): Data inicial para análise. Se None, usa a data da primeira operação.
            data_fim (datetime, optional): Data final para análise. Se None, usa a data atual.
            
        Returns:
            Dict: Dicionário com o relatório de desempenho.
        """
        try:
            # Define o período de análise
            data_fim = data_fim or datetime.utcnow()
            
            # Se não for fornecida data de início, usa a data da primeira operação
            if data_inicio is None:
                primeira_operacao = Operacao.query.filter_by(usuario_id=usuario_id)\
                    .order_by(Operacao.data_operacao.asc()).first()
                data_inicio = primeira_operacao.data_operacao if primeira_operacao else data_fim
            
            # Obtém o saldo inicial e final da carteira
            saldo_inicial = Decimal('0')
            saldo_final = Decimal('0')
            
            # Obtém todas as operações no período
            operacoes = Operacao.query.filter(
                Operacao.usuario_id == usuario_id,
                Operacao.data_operacao.between(data_inicio, data_fim)
            ).order_by(Operacao.data_operacao.asc()).all()
            
            # Calcula o saldo inicial (posições antes da data de início)
            posicoes_iniciais = db.session.query(
                Carteira.ativo_id,
                Carteira.quantidade,
                Carteira.preco_medio
            ).filter(
                Carteira.usuario_id == usuario_id
            ).all()
            
            # Ajusta as posições iniciais com base nas operações anteriores
            for operacao in Operacao.query.filter(
                Operacao.usuario_id == usuario_id,
                Operacao.data_operacao < data_inicio
            ).all():
                # Lógica para ajustar as posições iniciais com base nas operações anteriores
                # (implementação simplificada - pode ser expandida conforme necessário)
                pass
            
            # Calcula o valor inicial da carteira
            for posicao in posicoes_iniciais:
                saldo_inicial += posicao.quantidade * posicao.preco_medio
            
            # Processa as operações no período
            for operacao in operacoes:
                if operacao.tipo in [TipoOperacao.COMPRA, TipoOperacao.TRANSFERENCIA_ENTRADA, TipoOperacao.BONIFICACAO]:
                    saldo_final += operacao.valor_total if operacao.valor_total else Decimal('0')
                elif operacao.tipo in [TipoOperacao.VENDA, TipoOperacao.TRANSFERENCIA_SAIDA]:
                    saldo_final -= operacao.valor_total if operacao.valor_total else Decimal('0')
            
            # Obtém o saldo final atual
            posicoes_atuais = Carteira.query.filter_by(usuario_id=usuario_id).all()
            for posicao in posicoes_atuais:
                ativo = Ativo.query.get(posicao.ativo_id)
                if ativo and ativo.preco_atual:
                    saldo_final += posicao.quantidade * ativo.preco_atual
                else:
                    saldo_final += posicao.quantidade * posicao.preco_medio
            
            # Calcula o retorno absoluto e percentual
            retorno_absoluto = saldo_final - saldo_inicial
            retorno_percentual = (retorno_absoluto / saldo_inicial * 100) if saldo_inicio > 0 else Decimal('0')
            
            # Obtém estatísticas por ativo
            estatisticas_ativos = []
            posicoes = Carteira.query.filter_by(usuario_id=usuario_id).all()
            
            for posicao in posicoes:
                ativo = Ativo.query.get(posicao.ativo_id)
                if not ativo:
                    continue
                
                # Calcula o valor atual da posição
                preco_atual = ativo.preco_atual if ativo.preco_atual else posicao.preco_medio
                valor_atual = posicao.quantidade * preco_atual
                
                # Calcula o lucro/prejuízo
                valor_investido = posicao.quantidade * posicao.preco_medio
                lucro = valor_atual - valor_investido
                lucro_percentual = (lucro / valor_investido * 100) if valor_investido > 0 else Decimal('0')
                
                # Obtém o histórico de preços para análise de volatilidade
                # (implementação simplificada - pode ser expandida com dados históricos reais)
                volatilidade = Decimal('0')  # Placeholder para cálculo de volatilidade
                
                estatisticas_ativos.append({
                    'ativo_id': posicao.ativo_id,
                    'ticker': ativo.ticker,
                    'nome': ativo.nome,
                    'quantidade': float(posicao.quantidade),
                    'preco_medio': float(posicao.preco_medio),
                    'preco_atual': float(preco_atual),
                    'valor_investido': float(valor_investido),
                    'valor_atual': float(valor_atual),
                    'lucro_absoluto': float(lucro),
                    'lucro_percentual': float(lucro_percentual),
                    'volatilidade': float(volatilidade),
                    'percentual_carteira': float((valor_atual / saldo_final) * 100) if saldo_final > 0 else 0,
                    'setor': ativo.setor,
                    'subsetor': ativo.subsetor,
                    'segmento': ativo.segmento
                })
            
            # Ordena os ativos por valor atual (maior primeiro)
            estatisticas_ativos.sort(key=lambda x: x['valor_atual'], reverse=True)
            
            # Calcula estatísticas gerais
            total_investido = sum(p['valor_investido'] for p in estatisticas_ativos)
            total_atual = sum(p['valor_atual'] for p in estatisticas_ativos)
            lucro_total = total_atual - total_investido
            lucro_percentual_total = (lucro_total / total_investido * 100) if total_investido > 0 else 0
            
            # Agrupa por setor
            posicoes_por_setor = {}
            for ativo in estatisticas_ativos:
                setor = ativo['setor'] or 'Outros'
                if setor not in posicoes_por_setor:
                    posicoes_por_setor[setor] = {
                        'valor_total': Decimal('0'),
                        'percentual': 0,
                        'lucro_total': Decimal('0'),
                        'quantidade_ativos': 0
                    }
                
                posicoes_por_setor[setor]['valor_total'] += Decimal(str(ativo['valor_atual']))
                posicoes_por_setor[setor]['lucro_total'] += Decimal(str(ativo['lucro_absoluto']))
                posicoes_por_setor[setor]['quantidade_ativos'] += 1
            
            # Calcula percentuais
            for setor in posicoes_por_setor:
                posicoes_por_setor[setor]['percentual'] = \
                    float((posicoes_por_setor[setor]['valor_total'] / total_atual * 100) if total_atual > 0 else 0)
                posicoes_por_setor[setor]['lucro_total'] = float(posicoes_por_setor[setor]['lucro_total'])
                posicoes_por_setor[setor]['valor_total'] = float(posicoes_por_setor[setor]['valor_total'])
            
            # Ordena setores por valor total (maior primeiro)
            setores_ordenados = sorted(
                [{'setor': k, **v} for k, v in posicoes_por_setor.items()],
                key=lambda x: x['valor_total'],
                reverse=True
            )
            
            return {
                'periodo': {
                    'data_inicio': data_inicio.isoformat(),
                    'data_fim': data_fim.isoformat(),
                    'dias': (data_fim - data_inicio).days if data_inicio else 0
                },
                'saldo_inicial': float(saldo_inicial),
                'saldo_final': float(saldo_final),
                'retorno_absoluto': float(retorno_absoluto),
                'retorno_percentual': float(retorno_percentual),
                'estatisticas_gerais': {
                    'total_investido': float(total_investido),
                    'valor_atual': float(total_atual),
                    'lucro_absoluto': float(lucro_total),
                    'lucro_percentual': float(lucro_percentual_total),
                    'quantidade_ativos': len(estatisticas_ativos),
                    'ativos_lucro': len([a for a in estatisticas_ativos if a['lucro_absoluto'] >= 0]),
                    'ativos_prejuizo': len([a for a in estatisticas_ativos if a['lucro_absoluto'] < 0]),
                    'maior_lucro': max([a['lucro_absoluto'] for a in estatisticas_ativos], default=0),
                    'maior_prejuizo': min([a['lucro_absoluto'] for a in estatisticas_ativos], default=0)
                },
                'distribuicao_setorial': setores_ordenados,
                'detalhes_ativos': estatisticas_ativos,
                'data_geracao': datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório de desempenho: {str(e)}")
            raise
    
    @staticmethod
    def obter_historico_operacoes(
        usuario_id: int,
        ativo_id: Optional[int] = None,
        tipo_operacao: Optional[TipoOperacao] = None,
        data_inicio: Optional[datetime] = None,
        data_fim: Optional[datetime] = None,
        pagina: int = 1,
        itens_por_pagina: int = 50
    ) -> Dict:
        """
        Obtém o histórico de operações da carteira do usuário com filtros opcionais.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int, optional): ID do ativo para filtrar. Se None, retorna de todos os ativos.
            tipo_operacao (TipoOperacao, optional): Tipo de operação para filtrar. Se None, retorna todos os tipos.
            data_inicio (datetime, optional): Data inicial para filtrar. Se None, sem limite inferior.
            data_fim (datetime, optional): Data final para filtrar. Se None, sem limite superior.
            pagina (int, optional): Número da página para paginação. Padrão: 1.
            itens_por_pagina (int, optional): Quantidade de itens por página. Padrão: 50.
            
        Returns:
            Dict: Dicionário contendo a lista de operações e informações de paginação.
        """
        try:
            # Validação dos parâmetros
            if pagina < 1:
                pagina = 1
                
            if itens_por_pagina < 1 or itens_por_pagina > 1000:
                itens_por_pagina = 50
            
            # Inicia a consulta
            query = Operacao.query.filter_by(usuario_id=usuario_id)
            
            # Aplica filtros opcionais
            if ativo_id is not None:
                query = query.filter_by(ativo_id=ativo_id)
                
            if tipo_operacao is not None:
                query = query.filter_by(tipo=tipo_operacao)
                
            if data_inicio is not None:
                query = query.filter(Operacao.data_operacao >= data_inicio)
                
            if data_fim is not None:
                # Adiciona 1 dia para incluir todo o dia final
                data_fim = data_fim.replace(hour=23, minute=59, second=59, microsecond=999999)
                query = query.filter(Operacao.data_operacao <= data_fim)
            
            # Ordena por data de operação (mais recentes primeiro)
            query = query.order_by(Operacao.data_operacao.desc())
            
            # Calcula a paginação
            total_itens = query.count()
            total_paginas = (total_itens + itens_por_pagina - 1) // itens_por_pagina
            
            # Ajusta a página se necessário
            if pagina > total_paginas and total_paginas > 0:
                pagina = total_paginas
            
            # Aplica a paginação
            offset = (pagina - 1) * itens_por_pagina
            operacoes = query.offset(offset).limit(itens_por_pagina).all()
            
            # Formata os resultados
            operacoes_formatadas = [
                {
                    'id': op.id,
                    'tipo': op.tipo.name.lower(),
                    'ativo_id': op.ativo_id,
                    'quantidade': float(op.quantidade) if op.quantidade else None,
                    'preco_unitario': float(op.preco_unitario) if op.preco_unitario else None,
                    'valor_total': float(op.valor_total) if op.valor_total else None,
                    'valor_liquido': float(op.valor_liquido) if op.valor_liquido else None,
                    'custos': float(op.custos) if op.custos else 0,
                    'data_operacao': op.data_operacao.isoformat(),
                    'data_com': op.data_com.isoformat() if op.data_com else None,
                    'data_pagamento': op.data_pagamento.isoformat() if op.data_pagamento else None,
                    'status': op.status.name.lower(),
                    'notas': op.notas,
                    'criado_em': op.criado_em.isoformat() if op.criado_em else None,
                    'atualizado_em': op.atualizado_em.isoformat() if op.atualizado_em else None
                }
                for op in operacoes
            ]
            
            return {
                'operacoes': operacoes_formatadas,
                'paginacao': {
                    'pagina_atual': pagina,
                    'itens_por_pagina': itens_por_pagina,
                    'total_itens': total_itens,
                    'total_paginas': total_paginas,
                    'tem_proxima': pagina < total_paginas,
                    'tem_anterior': pagina > 1
                },
                'filtros': {
                    'ativo_id': ativo_id,
                    'tipo_operacao': tipo_operacao.name.lower() if tipo_operacao else None,
                    'data_inicio': data_inicio.isoformat() if data_inicio else None,
                    'data_fim': data_fim.isoformat() if data_fim else None
                }
            }
            
        except Exception as e:
            logger.error(f"Erro ao obter histórico de operações: {str(e)}")
            raise
    
    @staticmethod
    def registrar_bonificacao(
        usuario_id: int,
        ativo_id: int,
        quantidade: float,
        valor_unitario: float = 0,
        data_operacao: Optional[datetime] = None,
        notas: str = None
    ) -> Dict:
        """
        Registra o recebimento de bonificação (ações gratuitas) na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int): ID do ativo que está fornecendo a bonificação.
            quantidade (float): Quantidade de ativos recebidos como bonificação.
            valor_unitario (float, optional): Valor unitário para cálculo do preço médio. Se 0, não altera o preço médio.
            data_operacao (datetime, optional): Data da bonificação. Se None, usa a data/hora atual.
            notas (str, optional): Notas adicionais sobre a bonificação.
            
        Returns:
            Dict: Dicionário com os detalhes da bonificação e a nova posição.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado na carteira.
        """
        try:
            # Validação dos parâmetros
            if quantidade <= 0:
                raise ValidationError("A quantidade deve ser maior que zero.")
                
            if valor_unitario < 0:
                raise ValidationError("O valor unitário não pode ser negativo.")
            
            # Verifica se o ativo existe na carteira do usuário
            posicao = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            if not posicao:
                # Se não existir posição, cria uma nova
                posicao = Carteira(
                    usuario_id=usuario_id,
                    ativo_id=ativo_id,
                    quantidade=Decimal('0'),
                    preco_medio=Decimal('0'),
                    data_entrada=datetime.utcnow(),
                    data_atualizacao=datetime.utcnow()
                )
                db.session.add(posicao)
            
            # Salva a posição anterior
            posicao_anterior = {
                'quantidade': float(posicao.quantidade),
                'preco_medio': float(posicao.preco_medio)
            }
            
            # Converte para Decimal para evitar problemas de arredondamento
            quantidade = Decimal(str(quantidade))
            valor_unitario = Decimal(str(valor_unitario))
            
            # Calcula o novo preço médio se valor_unitario for maior que zero
            if valor_unitario > 0:
                valor_total_anterior = posicao.quantidade * posicao.preco_medio
                valor_total_bonus = quantidade * valor_unitario
                nova_quantidade = posicao.quantidade + quantidade
                
                if nova_quantidade > 0:
                    novo_preco_medio = (valor_total_anterior + valor_total_bonus) / nova_quantidade
                else:
                    novo_preco_medio = Decimal('0')
                
                posicao.preco_medio = novo_preco_medio
            
            # Atualiza a quantidade
            posicao.quantidade += quantidade
            posicao.data_atualizacao = datetime.utcnow()
            
            # Data da operação (usa a data atual se não fornecida)
            data_operacao = data_operacao or datetime.utcnow()
            
            # Registra a bonificação no histórico
            operacao = Operacao(
                usuario_id=usuario_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.BONIFICACAO,
                quantidade=quantidade,
                preco_unitario=valor_unitario,
                valor_total=quantidade * valor_unitario if valor_unitario > 0 else Decimal('0'),
                data_operacao=data_operacao,
                status=StatusOperacao.CONCLUIDA,
                notas=notas or "Bonificação de ações"
            )
            db.session.add(operacao)
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            return {
                'operacao': {
                    'id': operacao.id,
                    'tipo': 'bonificacao',
                    'quantidade': float(quantidade),
                    'valor_unitario': float(valor_unitario) if valor_unitario > 0 else None,
                    'data_operacao': data_operacao.isoformat(),
                    'notas': notas
                },
                'posicao_anterior': posicao_anterior,
                'posicao_atual': {
                    'quantidade': float(posicao.quantidade),
                    'preco_medio': float(posicao.preco_medio),
                    'valor_total': float(posicao.quantidade * posicao.preco_medio)
                }
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao registrar bonificação: {str(e)}")
            raise
    
    @staticmethod
    def registrar_desdobramento(
        usuario_id: int,
        ativo_id: int,
        proporcao: str,  # Formato: "5:1" (5 para 1)
        data_operacao: Optional[datetime] = None,
        notas: str = None
    ) -> Dict:
        """
        Registra um desdobramento (split) de ativo na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int): ID do ativo que sofreu o desdobramento.
            proporcao (str): Proporção do desdobramento no formato "X:Y" (ex: "5:1" para 5 por 1).
            data_operacao (datetime, optional): Data do desdobramento. Se None, usa a data/hora atual.
            notas (str, optional): Notas adicionais sobre o desdobramento.
            
        Returns:
            Dict: Dicionário com os detalhes do desdobramento e a nova posição.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado na carteira.
        """
        try:
            # Validação da proporção
            try:
                antes, depois = map(int, proporcao.split(':'))
                if antes <= 0 or depois <= 0:
                    raise ValueError("Os valores da proporção devem ser positivos.")
            except (ValueError, AttributeError):
                raise ValidationError("A proporção deve estar no formato 'X:Y' (ex: '5:1').")
                
            # Verifica se o ativo existe na carteira do usuário
            posicao = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            if not posicao:
                raise AssetNotFoundError(f"Ativo com ID {ativo_id} não encontrado na carteira.")
            
            # Data da operação (usa a data atual se não fornecida)
            data_operacao = data_operacao or datetime.utcnow()
            
            # Calcula a nova quantidade após o desdobramento
            fator = Decimal(str(antes)) / Decimal(str(depois))
            nova_quantidade = posicao.quantidade * fator
            
            # Atualiza a posição na carteira
            posicao_anterior = {
                'quantidade': float(posicao.quantidade),
                'preco_medio': float(posicao.preco_medio)
            }
            
            posicao.quantidade = nova_quantidade
            posicao.preco_medio = (posicao.quantidade * posicao.preco_medio) / nova_quantidade if nova_quantidade > 0 else Decimal('0')
            posicao.data_atualizacao = datetime.utcnow()
            
            # Registra o desdobramento no histórico
            operacao = Operacao(
                usuario_id=usuario_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.DESDOBRAMENTO,
                quantidade=nova_quantidade,
                preco_unitario=posicao.preco_medio,
                valor_total=Decimal('0'),  # Operação sem impacto financeiro direto
                data_operacao=data_operacao,
                status=StatusOperacao.CONCLUIDA,
                notas=f"Desdobramento {proporcao}" + (f" - {notas}" if notas else "")
            )
            db.session.add(operacao)
            
            # Atualiza o preço atual do ativo (ajustado pelo split)
            ativo = Ativo.query.get(ativo_id)
            if ativo and ativo.preco_atual:
                ativo.preco_atual = ativo.preco_atual * (Decimal(str(depois)) / Decimal(str(antes)))
                ativo.ultima_atualizacao = datetime.utcnow()
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            return {
                'operacao': {
                    'id': operacao.id,
                    'tipo': 'desdobramento',
                    'proporcao': proporcao,
                    'fator': float(fator),
                    'data_operacao': data_operacao.isoformat(),
                    'notas': notas
                },
                'posicao_anterior': posicao_anterior,
                'posicao_atual': {
                    'quantidade': float(posicao.quantidade),
                    'preco_medio': float(posicao.preco_medio),
                    'valor_total': float(posicao.quantidade * posicao.preco_medio)
                }
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao registrar desdobramento: {str(e)}")
            raise
    
    @staticmethod
    def registrar_venda(
        usuario_id: int,
        ativo_id: int,
        quantidade: float,
        preco_unitario: float,
        data_operacao: Optional[datetime] = None,
        custos: float = 0,
        notas: str = None,
        vender_tudo: bool = False
    ) -> Dict:
        """
        Registra uma operação de venda de ativo na carteira do usuário.
        
        Args:
            usuario_id (int): ID do usuário.
            ativo_id (int): ID do ativo sendo vendido.
            quantidade (float): Quantidade de ativos vendidos. Se vender_tudo=True, este valor é ignorado.
            preco_unitario (float): Preço unitário recebido por cada ativo.
            data_operacao (datetime, optional): Data da operação. Se None, usa a data/hora atual.
            custos (float, optional): Custos adicionais da operação. Defaults to 0.
            notas (str, optional): Notas adicionais sobre a operação.
            vender_tudo (bool, optional): Se True, vende toda a posição do ativo. Defaults to False.
            
        Returns:
            Dict: Dicionário com os detalhes da operação e o novo saldo.
            
        Raises:
            ValidationError: Se os parâmetros forem inválidos.
            AssetNotFoundError: Se o ativo não for encontrado na carteira.
            InsufficientFundsError: Se não houver saldo suficiente para a venda.
        """
        try:
            # Validação dos parâmetros
            if not vender_tudo and quantidade <= 0:
                raise ValidationError("A quantidade deve ser maior que zero.")
                
            if preco_unitario < 0:
                raise ValidationError("O preço unitário não pode ser negativo.")
                
            if custos < 0:
                raise ValidationError("Os custos não podem ser negativos.")
            
            # Verifica se o ativo existe na carteira do usuário
            posicao = Carteira.query.filter_by(
                usuario_id=usuario_id,
                ativo_id=ativo_id
            ).first()
            
            if not posicao or posicao.quantidade <= 0:
                raise AssetNotFoundError(f"Ativo com ID {ativo_id} não encontrado na carteira ou sem saldo disponível.")
            
            # Converte para Decimal para evitar problemas de arredondamento
            quantidade_disponivel = posicao.quantidade
            
            if vender_tudo:
                quantidade_venda = quantidade_disponivel
            else:
                quantidade_venda = Decimal(str(quantidade))
                if quantidade_venda > quantidade_disponivel:
                    raise InsufficientFundsError(
                        f"Quantidade insuficiente para venda. Disponível: {quantidade_disponivel}"
                    )
            
            preco_unitario = Decimal(str(preco_unitario))
            custos = Decimal(str(custos))
            
            # Calcula o valor total da operação (valor bruto - custos)
            valor_bruto = quantidade_venda * preco_unitario
            valor_liquido = valor_bruto - custos
            
            # Data da operação (usa a data atual se não fornecida)
            data_operacao = data_operacao or datetime.utcnow()
            
            # Atualiza a posição na carteira
            nova_quantidade = quantidade_disponivel - quantidade_venda
            
            if nova_quantidade <= 0:
                # Remove a posição se o saldo zerar
                db.session.delete(posicao)
            else:
                # Atualiza a quantidade mantendo o preço médio
                posicao.quantidade = nova_quantidade
                posicao.data_atualizacao = datetime.utcnow()
            
            # Registra a operação no histórico
            operacao = Operacao(
                usuario_id=usuario_id,
                ativo_id=ativo_id,
                tipo=TipoOperacao.VENDA,
                quantidade=quantidade_venda,
                preco_unitario=preco_unitario,
                valor_total=valor_bruto,
                valor_liquido=valor_liquido,
                custos=custos,
                data_operacao=data_operacao,
                status=StatusOperacao.CONCLUIDA,
                notas=notas
            )
            db.session.add(operacao)
            
            # Atualiza o preço atual do ativo
            ativo = Ativo.query.get(ativo_id)
            if ativo:
                ativo.preco_atual = preco_unitario
                ativo.ultima_atualizacao = datetime.utcnow()
            
            # Salva as alterações no banco de dados
            db.session.commit()
            
            # Verifica se ainda há posição após a venda
            posicao_atualizada = None
            if nova_quantidade > 0:
                posicao_atualizada = Carteira.query.filter_by(
                    usuario_id=usuario_id,
                    ativo_id=ativo_id
                ).first()
            
            return {
                'operacao': {
                    'id': operacao.id,
                    'tipo': 'venda',
                    'quantidade': float(quantidade_venda),
                    'preco_unitario': float(preco_unitario),
                    'valor_bruto': float(valor_bruto),
                    'valor_liquido': float(valor_liquido),
                    'custos': float(custos),
                    'data_operacao': data_operacao.isoformat(),
                    'notas': notas
                },
                'posicao_anterior': {
                    'quantidade': float(quantidade_disponivel),
                    'preco_medio': float(posicao.preco_medio)
                },
                'posicao_atual': {
                    'quantidade': float(posicao_atualizada.quantidade) if posicao_atualizada else 0,
                    'preco_medio': float(posicao_atualizada.preco_medio) if posicao_atualizada else 0,
                    'valor_total': float(posicao_atualizada.quantidade * posicao_atualizada.preco_medio) if posicao_atualizada else 0
                } if posicao_atualizada else None
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao registrar venda: {str(e)}")
            raise

# Exemplo de uso:
if __name__ == "__main__":
    from .. import create_app
    app = create_app()
    
    with app.app_context():
        try:
            # Exemplo de como usar o serviço
            carteira_service = CarteiraService()
            carteira = carteira_service.obter_carteira(usuario_id=1)
            print(f"Carteira: {carteira}")
        except Exception as e:
            print(f"Erro: {str(e)}")
