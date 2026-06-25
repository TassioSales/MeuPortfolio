"""Export expenses to CSV and XLSX."""
import csv
import io
import os
from datetime import datetime

from loguru import logger


def _cabecalho() -> list[str]:
    return ["ID", "Data", "Descrição", "Categoria", "Valor (R$)"]


def exportar_csv(gastos: list[dict], mes: str) -> bytes:
    """Return UTF-8 CSV bytes for the given expenses."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_cabecalho())
    for g in gastos:
        writer.writerow([
            g.get("id", ""),
            g.get("data", ""),
            g.get("descricao", ""),
            g.get("categoria", ""),
            f"{g.get('valor', 0):.2f}",
        ])
    total = sum(g.get("valor", 0) for g in gastos)
    writer.writerow(["", "", "", "TOTAL", f"{total:.2f}"])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def exportar_xlsx(gastos: list[dict], mes: str) -> bytes:
    """Return XLSX bytes with formatted spreadsheet."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl não instalado")
        raise

    from bot.services.formatters import MESES_NOME, emoji_categoria

    wb = openpyxl.Workbook()
    ws = wb.active

    try:
        ano, num = mes.split("-")
        nome_mes = f"{MESES_NOME.get(num, num)} {ano}"
    except ValueError:
        nome_mes = mes

    ws.title = f"Gastos {nome_mes}"

    # Colors
    verde_escuro = "1B5E20"
    verde_claro = "E8F5E9"
    cinza = "F5F5F5"
    borda_cor = "C8E6C9"

    borda = Border(
        left=Side(style="thin", color=borda_cor),
        right=Side(style="thin", color=borda_cor),
        top=Side(style="thin", color=borda_cor),
        bottom=Side(style="thin", color=borda_cor),
    )

    # Title row
    ws.merge_cells("A1:E1")
    titulo = ws["A1"]
    titulo.value = f"💰 Relatório de Gastos — {nome_mes}"
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=verde_escuro)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font = Font(italic=True, size=9, color="666666")
    sub.alignment = Alignment(horizontal="center")

    # Header row
    headers = _cabecalho()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
    ws.row_dimensions[3].height = 20

    # Data rows
    for row_idx, g in enumerate(gastos, 4):
        fill = PatternFill("solid", fgColor=verde_claro if row_idx % 2 == 0 else "FFFFFF")
        values = [
            g.get("id", ""),
            g.get("data", ""),
            g.get("descricao", ""),
            g.get("categoria", "").capitalize(),
            g.get("valor", 0),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = borda
            cell.alignment = Alignment(vertical="center")
            if col == 5:  # Valor
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col == 1:  # ID
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Total row
    total_row = len(gastos) + 4
    total = sum(g.get("valor", 0) for g in gastos)

    ws.merge_cells(f"A{total_row}:D{total_row}")
    label = ws.cell(row=total_row, column=1, value="TOTAL GERAL")
    label.font = Font(bold=True, color="FFFFFF")
    label.fill = PatternFill("solid", fgColor=verde_escuro)
    label.alignment = Alignment(horizontal="right", vertical="center")
    label.border = borda

    total_cell = ws.cell(row=total_row, column=5, value=total)
    total_cell.font = Font(bold=True, color="FFFFFF")
    total_cell.fill = PatternFill("solid", fgColor=verde_escuro)
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = borda
    ws.row_dimensions[total_row].height = 22

    # Category summary sheet
    ws2 = wb.create_sheet("Por Categoria")
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value = f"Gastos por Categoria — {nome_mes}"
    t2.font = Font(bold=True, size=12, color="FFFFFF")
    t2.fill = PatternFill("solid", fgColor=verde_escuro)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    cats: dict[str, dict] = {}
    for g in gastos:
        cat = g.get("categoria", "outros").capitalize()
        if cat not in cats:
            cats[cat] = {"qtd": 0, "total": 0.0}
        cats[cat]["qtd"] += 1
        cats[cat]["total"] += g.get("valor", 0)

    for col, h in enumerate(["Categoria", "Qtd.", "Total (R$)"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center")
        cell.border = borda

    for i, (cat, info) in enumerate(sorted(cats.items(), key=lambda x: -x[1]["total"]), 3):
        fill = PatternFill("solid", fgColor=verde_claro if i % 2 == 0 else "FFFFFF")
        for col, val in enumerate([cat, info["qtd"], info["total"]], 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.border = borda
            if col == 3:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(f"XLSX gerado: {len(gastos)} gastos, mês {mes}")
    return buf.read()


def salvar_em_disco(dados: bytes, nome_arquivo: str, pasta: str = "data/exports") -> str:
    """Save bytes to disk and return the full path."""
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, nome_arquivo)
    with open(caminho, "wb") as f:
        f.write(dados)
    return caminho
