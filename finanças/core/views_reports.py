"""Reports, CSV export, PDF export, and Excel export views."""
import csv
import io
import json
import logging
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.db.models.functions import TruncDay, TruncMonth
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.utils import timezone
from xhtml2pdf import pisa

from .models import Budget, Category, Transaction

logger = logging.getLogger("core")


def _filter_transactions(request, start_date, end_date, category_id):
    qs = Transaction.objects.filter(user=request.user).order_by("-date")
    if start_date:
        qs = qs.filter(date__gte=start_date)
    if end_date:
        qs = qs.filter(date__lte=end_date)
    if category_id:
        qs = qs.filter(category_id=category_id)
    return qs


@login_required
def reports(request):
    logger.info(f"Generating reports for user {request.user.username}")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    category_id = request.GET.get("category")

    transactions = _filter_transactions(request, start_date, end_date, category_id)

    total_income = (
        transactions.filter(type="RECEITA").aggregate(Sum("amount"))["amount__sum"] or 0
    )
    total_expense = (
        transactions.filter(type="DESPESA").aggregate(Sum("amount"))["amount__sum"] or 0
    )
    net_balance = total_income - total_expense
    savings_rate = (
        ((total_income - total_expense) / total_income) * 100 if total_income > 0 else 0
    )

    categories = Category.objects.filter(user=request.user)

    expense_by_category = (
        transactions.filter(type="DESPESA")
        .values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )

    evolution_qs = Transaction.objects.filter(user=request.user)
    if start_date:
        evolution_qs = evolution_qs.filter(date__gte=start_date)
    if end_date:
        evolution_qs = evolution_qs.filter(date__lte=end_date)
    if not start_date and not end_date:
        last_year = timezone.now().date() - timedelta(days=365)
        evolution_qs = evolution_qs.filter(date__gte=last_year)

    monthly_data = (
        evolution_qs.annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(
            income=Sum("amount", filter=Q(type="RECEITA")),
            expense=Sum("amount", filter=Q(type="DESPESA")),
        )
        .order_by("month")
    )

    evolution_labels = []
    evolution_income = []
    evolution_expense = []
    for entry in monthly_data:
        if entry["month"]:
            evolution_labels.append(entry["month"].strftime("%b/%Y"))
            evolution_income.append(float(entry["income"] or 0))
            evolution_expense.append(float(entry["expense"] or 0))

    daily_data = (
        transactions.filter(type="DESPESA")
        .annotate(day=TruncDay("date"))
        .values("day")
        .annotate(total=Sum("amount"))
        .order_by("day")
    )
    daily_labels = []
    daily_expenses = []
    for entry in daily_data:
        if entry["day"]:
            daily_labels.append(entry["day"].strftime("%d/%m"))
            daily_expenses.append(float(entry["total"] or 0))

    budgets = Budget.objects.filter(user=request.user)
    budget_labels = []
    budget_limits = []
    budget_actuals = []
    for budget in budgets:
        actual = (
            transactions.filter(
                category=budget.category, type="DESPESA"
            ).aggregate(Sum("amount"))["amount__sum"]
            or 0
        )
        budget_labels.append(budget.category.name)
        budget_limits.append(float(budget.limit))
        budget_actuals.append(float(actual))

    context = {
        "categories": categories,
        "total_income": total_income,
        "total_expense": total_expense,
        "net_balance": net_balance,
        "savings_rate": savings_rate,
        "expense_by_category": expense_by_category,
        "top_expenses": transactions.filter(type="DESPESA").order_by("-amount")[:5],
        "evolution_labels": json.dumps(evolution_labels),
        "evolution_income": json.dumps(evolution_income),
        "evolution_expense": json.dumps(evolution_expense),
        "daily_labels": json.dumps(daily_labels),
        "daily_expenses": json.dumps(daily_expenses),
        "budget_labels": json.dumps(budget_labels),
        "budget_limits": json.dumps(budget_limits),
        "budget_actuals": json.dumps(budget_actuals),
        "recent_transactions": transactions.order_by("-date", "-id")[:20],
    }
    return render(request, "core/reports.html", context)


@login_required
def export_csv(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    category_id = request.GET.get("category")

    transactions = _filter_transactions(request, start_date, end_date, category_id)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="transactions.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Type", "Category", "Amount", "Description"])
    for t in transactions:
        writer.writerow([
            t.date,
            t.get_type_display(),
            t.category.name if t.category else "-",
            t.amount,
            t.description,
        ])
    return response


@login_required
def export_pdf(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    category_id = request.GET.get("category")

    transactions = _filter_transactions(request, start_date, end_date, category_id)

    total_income = (
        transactions.filter(type="RECEITA").aggregate(Sum("amount"))["amount__sum"] or 0
    )
    total_expense = (
        transactions.filter(type="DESPESA").aggregate(Sum("amount"))["amount__sum"] or 0
    )

    context = {
        "transactions": transactions,
        "total_income": total_income,
        "total_expense": total_expense,
        "net_balance": total_income - total_expense,
        "start_date": start_date,
        "end_date": end_date,
    }

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="relatorio_financeiro.pdf"'

    template = get_template("core/reports_pdf.html")
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("We had some errors <pre>" + html + "</pre>")
    return response


@login_required
def export_xlsx(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    category_id = request.GET.get("category")

    transactions = _filter_transactions(request, start_date, end_date, category_id)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Transactions ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transações"

    header_fill = PatternFill("solid", fgColor="1a73e8")
    header_font = Font(bold=True, color="FFFFFF")
    income_fill = PatternFill("solid", fgColor="d4edda")
    expense_fill = PatternFill("solid", fgColor="f8d7da")

    headers = ["Data", "Tipo", "Categoria", "Valor (R$)", "Descrição", "Método de Pagamento"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [12, 10, 20, 14, 40, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_idx, t in enumerate(transactions, 2):
        row_fill = income_fill if t.type == "RECEITA" else expense_fill
        values = [
            t.date,
            t.get_type_display(),
            t.category.name if t.category else "-",
            float(t.amount),
            t.description,
            t.get_payment_method_display() if hasattr(t, "get_payment_method_display") else "-",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            if col == 4:
                cell.number_format = '#,##0.00'

    # Totals row
    total_row = len(list(transactions)) + 2
    total_income = transactions.filter(type="RECEITA").aggregate(Sum("amount"))["amount__sum"] or 0
    total_expense = transactions.filter(type="DESPESA").aggregate(Sum("amount"))["amount__sum"] or 0

    ws.cell(row=total_row, column=3, value="TOTAL RECEITAS").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=float(total_income)).number_format = '#,##0.00'
    ws.cell(row=total_row + 1, column=3, value="TOTAL DESPESAS").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=4, value=float(total_expense)).number_format = '#,##0.00'
    ws.cell(row=total_row + 2, column=3, value="SALDO LÍQUIDO").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=4, value=float(total_income - total_expense)).number_format = '#,##0.00'

    # ── Sheet 2: Resumo Mensal ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo Mensal")
    ws2.append(["Mês", "Receitas (R$)", "Despesas (R$)", "Saldo (R$)"])
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    from django.db.models.functions import TruncMonth as TM
    monthly = (
        transactions.annotate(mes=TM("date"))
        .values("mes")
        .annotate(
            receitas=Sum("amount", filter=Q(type="RECEITA")),
            despesas=Sum("amount", filter=Q(type="DESPESA")),
        )
        .order_by("mes")
    )
    for r, m in enumerate(monthly, 2):
        rec = float(m["receitas"] or 0)
        desp = float(m["despesas"] or 0)
        ws2.append([
            m["mes"].strftime("%b/%Y") if m["mes"] else "-",
            rec,
            desp,
            rec - desp,
        ])
        for col in [2, 3, 4]:
            ws2.cell(row=r, column=col).number_format = '#,##0.00'

    for col_w, width in zip(["A", "B", "C", "D"], [12, 16, 16, 14]):
        ws2.column_dimensions[col_w].width = width

    # ── Sheet 3: Por Categoria ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Categoria")
    ws3.append(["Categoria", "Total Despesas (R$)", "% do Total"])
    for col in range(1, 4):
        cell = ws3.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    cat_data = (
        transactions.filter(type="DESPESA")
        .values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )
    for r, c in enumerate(cat_data, 2):
        pct = float(c["total"] / total_expense * 100) if total_expense else 0
        ws3.append([c["category__name"] or "Sem categoria", float(c["total"]), pct])
        ws3.cell(row=r, column=2).number_format = '#,##0.00'
        ws3.cell(row=r, column=3).number_format = '0.0"%"'

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 14

    # ── Response ───────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="relatorio_financeiro.xlsx"'
    return response
